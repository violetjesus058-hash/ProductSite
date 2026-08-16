from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path('/home/ubuntu/product-catalog-site')
WORKBOOK = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
OUTPUT = ROOT / 'scripts/kakobuy-category-audit.md'

KEYWORDS = {
    'pants': r'\b(pants|trousers|jeans|shorts|joggers|sweatpants|leggings|denim|cargo|底裤|长裤|短裤|牛仔裤|休闲裤|运动裤)\b',
    'shoes': r'\b(shoes|shoe|sneakers|sneaker|boots|boot|slides|sandals|loafer|mule|running shoes)\b',
    'watches': r'\b(watch|watches|腕表|手表)\b',
    'fragrance': r'\b(perfume|fragrance|cologne|parfum|香水)\b',
    'bags': r'\b(bag|bags|backpack|tote|shoulder bag|crossbody|pouch|handbag)\b',
    'wallets': r'\b(wallet|card holder|cardholder|零钱包|钱包)\b',
    'clothing': r'\b(shirt|t-shirt|tee|hoodie|jacket|coat|sweater|sweatshirt|jersey|tracksuit|dress|skirt|vest|top|clothing|衬衫|卫衣|夹克|外套|毛衣|短袖|长袖|裙)\b',
    'accessories': r'\b(cap|hat|belt|glasses|sunglasses|scarf|tie|bracelet|ring|necklace|earring|accessory|case|headphones|phone)\b',
}

def clean(value: object) -> str:
    return re.sub(r'\s+', ' ', str(value or '').strip())

def main() -> None:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb['products']
    rows = ws.iter_rows(values_only=True)
    headers = [clean(v) for v in next(rows)]
    index = {name: i for i, name in enumerate(headers)}
    records = []
    for values in rows:
        title = clean(values[index.get('title_en_platform', -1)] if index.get('title_en_platform') is not None else '')
        original = clean(values[index.get('title_original', -1)] if index.get('title_original') is not None else '')
        category = clean(values[index.get('category', -1)] if index.get('category') is not None else '')
        subcategory = clean(values[index.get('subcategory', -1)] if index.get('subcategory') is not None else '')
        pid = clean(values[index.get('product_id', -1)] if index.get('product_id') is not None else '')
        text = f'{title} {original} {subcategory}'.lower()
        detected = [name for name, pattern in KEYWORDS.items() if re.search(pattern, text, re.I)]
        records.append({'pid': pid, 'title': title or original, 'category': category, 'subcategory': subcategory, 'detected': detected})
    raw_counts = Counter(item['category'] or 'Unclassified' for item in records)
    sub_counts = Counter((item['category'] or 'Unclassified', item['subcategory'] or 'Unspecified') for item in records)
    conflicts = [item for item in records if item['detected'] and item['category'] in {'Accessories', 'Unclassified', ''}]
    pants_conflicts = [item for item in conflicts if 'pants' in item['detected']]
    lines = ['# Kakobuy 分类审计', '', f'商品数：{len(records)}', '', '## 原始分类', '', '| 原始分类 | 商品数 |', '|---|---:|']
    lines.extend(f'| {key} | {value} |' for key, value in raw_counts.most_common())
    lines += ['', '## 关键词识别结果', '', '| 识别类型 | 商品数 |', '|---|---:|']
    detected_counts = Counter(name for item in records for name in item['detected'])
    lines.extend(f'| {key} | {value} |' for key, value in detected_counts.most_common())
    lines += ['', f'## 当前疑似错误：被归入 Accessories/Unclassified 但标题命中裤子关键词（{len(pants_conflicts)} 个）', '', '| Product ID | Title | Raw category | Subcategory |', '|---|---|---|---|']
    for item in pants_conflicts[:200]:
        lines.append(f"| {item['pid']} | {item['title'][:100].replace('|','/')} | {item['category']} | {item['subcategory']} |")
    lines += ['', '## 分类冲突样例（前 120 条）', '', '| Product ID | Title | Raw category | Detected |', '|---|---|---|---|']
    for item in conflicts[:120]:
        lines.append(f"| {item['pid']} | {item['title'][:100].replace('|','/')} | {item['category']} | {', '.join(item['detected'])} |")
    OUTPUT.write_text('\n'.join(lines) + '\n', encoding='utf-8')
    print(f'products={len(records)} raw_categories={dict(raw_counts)} detected={dict(detected_counts)} pants_conflicts={len(pants_conflicts)}')

if __name__ == '__main__':
    main()
