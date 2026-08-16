from __future__ import annotations

import json
import re
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

WORKBOOK = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
DOWNLOAD_REPORT = Path('/home/ubuntu/product-catalog-site/scripts/product-image-download-report.json')
LOCAL_IMAGE_DIR = Path('/home/ubuntu/product-catalog-site/client/public/product-images')
OUTPUT = Path('/home/ubuntu/product-catalog-site/scripts/kakobuy-cover-download-list.json')

def canon(url: object) -> str:
    if not url:
        return ''
    return re.sub(r'\?.*$', '', str(url).strip()).lower()

def main() -> None:
    download = json.loads(DOWNLOAD_REPORT.read_text(encoding='utf-8'))
    existing = {}
    for item in download.get('results', []):
        if item.get('ok') and item.get('url') and item.get('path'):
            stem = Path(item['path']).stem
            candidates = sorted(LOCAL_IMAGE_DIR.glob(f'{stem}_*.webp'))
            if candidates:
                existing[canon(item['url'])] = f'/product-images/{candidates[0].name}'
    chosen = {}
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb['product_images']
    rows = ws.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else '' for v in next(rows)]
    index = {name: i for i, name in enumerate(headers)}
    for values in rows:
        pid = values[index['product_id']]
        url = values[index['image_url']]
        if pid is None or not url:
            continue
        pid = str(pid).strip()
        item = {
            'product_id': pid,
            'url': str(url).strip(),
            'image_type': str(values[index['image_type']] or '').strip(),
            'image_order': values[index['image_order']] or 999,
            'local_path': existing.get(canon(url)),
        }
        current = chosen.get(pid)
        rank = (0 if item['image_type'].lower() == 'cover' else 1, item['image_order'])
        current_rank = (0 if current and current['image_type'].lower() == 'cover' else 1, current['image_order'] if current else 999999)
        if current is None or rank < current_rank:
            chosen[pid] = item
    output = list(chosen.values())
    missing = [item for item in output if not item['local_path']]
    OUTPUT.write_text(json.dumps({'generatedAt': '2026-08-16', 'products': len(output), 'reused': len(output)-len(missing), 'missing': len(missing), 'items': output}, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    print(f'Products: {len(output)}; reused covers: {len(output)-len(missing)}; missing covers to download: {len(missing)}')

if __name__ == '__main__':
    main()
