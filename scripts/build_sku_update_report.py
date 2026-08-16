from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
PRODUCTS = Path('/home/ubuntu/product-catalog-site/client/src/data/products.ts')
REPORT = Path('/home/ubuntu/product-catalog-site/scripts/sku-update-recommendation.md')


def norm(value: object) -> str:
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip().lower()


def pick(record: dict, names: list[str]) -> object:
    by_key = {norm(key): value for key, value in record.items()}
    for name in names:
        if norm(name) in by_key:
            return by_key[norm(name)]
    return None


def read_current() -> list[dict]:
    text = PRODUCTS.read_text(encoding='utf-8')
    match = re.search(r'export const products: Product\[\] = (\[.*?\]) as Product\[\];', text, re.S)
    return json.loads(match.group(1)) if match else []


def main() -> None:
    current = read_current()
    current_by_title: defaultdict[str, list[dict]] = defaultdict(list)
    for product in current:
        for field in ('name', 'catalogName'):
            title = norm(product.get(field))
            if title:
                current_by_title[title].append(product)

    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    product_rows: dict[str, dict] = {}
    sku_counts = Counter()
    image_counts = Counter()
    platform_counts = Counter()
    category_counts = Counter()
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration:
            continue
        headers = [str(value).strip() if value is not None else '' for value in header_values]
        for values in rows:
            record = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}
            product_id = pick(record, ['product_id'])
            if product_id is None:
                continue
            pid = str(product_id).strip()
            if ws.title == 'sku_records':
                sku_counts[pid] += 1
                if pid not in product_rows:
                    product_rows[pid] = record
            elif ws.title == 'product_images':
                image_counts[pid] += 1
            elif ws.title == 'product_platforms':
                platform_counts[pid] += 1
            elif ws.title == 'products':
                product_rows[pid] = record
                category_counts[str(pick(record, ['category']) or 'Unclassified')] += 1

    exact_matches: dict[str, dict] = {}
    for pid, record in product_rows.items():
        candidates = [pick(record, ['title_original']), pick(record, ['title_en_platform'])]
        matches = []
        for candidate in candidates:
            matches.extend(current_by_title.get(norm(candidate), []))
        unique = {item['id']: item for item in matches if item.get('id')}
        if unique:
            exact_matches[pid] = {
                'new': record,
                'current': list(unique.values()),
            }

    categories = Counter(str(pick(row, ['category']) or 'Unclassified') for row in product_rows.values())
    source_platforms = Counter(str(pick(row, ['primary_platform']) or '') for row in product_rows.values())
    report = [
        '# 最新 SKU 数据表更新建议', '',
        '## 结论摘要', '',
        '这份工作簿不是“每行一个独立商品”的简单商品表，而是一个以 `product_id` 为商品主键、以 `sku_id` 为选项明细的多表采集结果。当前网站的 276 个商品使用 Fansbuy micro 商品 ID；工作簿中的 1,154 个 `product_id` 是另一批商品的来源商品 ID，因此不能按 ID 直接覆盖，也不能把每一行 SKU 直接导入为商品卡片。', '',
        '| 项目 | 当前网站 | 最新工作簿 | 含义 |',
        '|---|---:|---:|---|',
        f'| 商品卡片/产品对象 | {len(current)} | {len(product_rows)} | 新表规模约为当前目录的 {len(product_rows)/max(1,len(current)):.1f} 倍 |',
        f'| SKU 明细 | — | {sum(sku_counts.values())} | 同一商品下存在多个款式、尺码或颜色组合 |',
        f'| 图片记录 | 已托管 | {sum(image_counts.values())} | 可用于新商品导入，但需先检查稳定性和联系方式风险 |',
        f'| 平台链接 | 每商品主入口为 Fansbuy | {sum(platform_counts.values())} | 新表主平台字段实际为 Kakobuy |',
        f'| 直接 product ID 重合 | — | 0 | 不能使用 product_id 直接匹配当前 276 项 |',
        f'| 标题精确匹配到当前目录 | — | {len(exact_matches)} | 可作为候选合并依据，但仍需 URL/图片二次确认 |',
        '',
        '## 数据结构判断', '',
        '工作簿包含 `sku_records`、`products`、`product_images`、`product_platforms`、`failed_pages`、`url_dedup` 和 `summary` 七类工作表。主表 `products` 提供商品级默认价格和库存汇总；`sku_records` 提供每个 SKU 的人民币价格、美元价格、库存和规格；`product_images` 提供主图与详情图；`product_platforms` 提供平台入口和参考价格。', '',
        '| 数据层 | 应在网站中的角色 | 更新建议 |',
        '|---|---|---|',
        '| products | 商品卡片和详情页的主记录 | 只生成一个商品对象，不按 SKU 行复制卡片 |',
        '| sku_records | 颜色、款式、尺码、库存和 SKU 价格 | 按商品聚合到 `sizes`、`colors` 或新的结构化 `variants` |',
        '| product_images | 主图、详情图、图片顺序 | 优先复用已托管图片；新图先下载、去重、清理敏感联系方式 |',
        '| product_platforms | 多平台购买入口 | 当前站点 Fansbuy 为主入口；Kakobuy 可作为新表来源链接保留 |',
        '| failed_pages | 失败采集记录 | 不导入前台，只用于数据质量报告 |',
        '',
        '## 重合与身份问题', '',
        f'虽然两个数据集的 ID 没有直接重合，但报告中的标题匹配结果显示至少存在一批同款候选。例如新表的“高品质短”能够匹配当前目录中的同名商品，但其 product_id 不同。这说明当前网站的 Fansbuy ID 与新表的 Weidian/Kakobuy 来源 ID 是不同层级的身份标识。正确做法是增加一个独立的来源映射，而不是修改现有 `id`。', '',
        '| 身份字段 | 建议保留方式 |',
        '|---|---|',
        '| `id` | 保持当前 Fansbuy micro ID，作为网站内部稳定路由 ID |',
        '| 新表 `product_id` | 增加 `sourceProductId` 或映射表字段 |',
        '| 新表 `sku_id` | 作为 `variants[].skuId`，不能替代商品 ID |',
        '| `source_url` | 作为原始 Weidian 链接保存 |',
        '| `platform_url` | 作为 Kakobuy 入口；如后续生成 Fansbuy 链接，再单独写入 `url` |',
        '',
        '## 推荐的三阶段更新方案', '',
        '### 阶段一：只做影子导入，不改变前台', '',
        '先将工作簿转换为一个独立的规范化数据文件和导入报告，建立 `sourceProductId → currentProductId` 的映射。对于标题精确匹配但 ID 不同的记录，必须同时比较原始商品链接、主图指纹、分类和价格；任何一项冲突都进入人工复核，不自动合并。', '',
        '### 阶段二：更新重合商品的价格和 SKU', '',
        '对于确认是同一商品的记录，只更新 SKU 层价格和库存，并用 `price_usd` 作为前台价格来源。商品级 `default_price_usd` 应取当前可售 SKU 的最低价或明确的默认 SKU，不能把一个随机 SKU 的价格写成整件商品的固定价格。价格字段必须同时保留 `price_rmb`、`price_usd`、`currency_source` 和 `price_checked_at`，以便后续继续审计。', '',
        '### 阶段三：将新商品作为增量目录加入', '',
        '对于工作簿中新出现且通过质量检查的商品，生成新的内部网站 ID，并把工作簿的 `product_id` 放在 `sourceProductId`。新增商品应先完成英文标题、分类、品牌判断、图片托管、Fansbuy 主链接和 SKU 聚合，再进入前台目录。', '',
        '## 我建议本次不要直接覆盖的原因', '',
        '第一，工作簿实际主平台是 Kakobuy，而当前网站要求 Fansbuy 为主入口；直接导入会造成入口平台和价格来源不一致。第二，工作簿有 57.9 万条失败页面记录中的 579 条失败记录，且产品、图片、平台和 SKU 数据需要按主键关联；直接覆盖可能把缺图、缺链接或失败记录带入前台。第三，当前项目的 `products.ts` 目前存在 Vite 报告的 `Unexpected "]"` 语法错误，必须先修复并建立可回滚快照，再进行大批量数据变更。', '',
        '## 推荐的实施顺序', '',
        '1. 先修复并验证 `products.ts` 当前语法错误。',
        '2. 生成规范化中间文件：商品表、SKU 表、图片表、平台表和身份映射表。',
        '3. 先导入 5–10 个样本商品，验证商品卡片、详情页、逐图购买模块、规格选择和 Fansbuy 链接。',
        '4. 人工确认标题相同但 ID 不同的合并候选，再批量更新价格和 SKU。',
        '5. 对剩余新商品执行增量导入，保留失败清单和回滚快照。',
        '6. 完成 TypeScript、生产构建和桌面/移动端抽样验证后，再创建项目检查点。',
        '',
        '## 当前建议的决策', '',
        '> 建议采用“保留当前 276 个 Fansbuy 商品 + 新表作为第二批候选商品池 + 通过映射后逐步合并”的策略，而不是用 1,154 个 Kakobuy 来源商品直接替换当前目录。', '',
        '本报告依据本地工作簿 `pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx`、现有 `client/src/data/products.ts` 和分析脚本生成。',
    ]
    REPORT.write_text('\n'.join(report) + '\n', encoding='utf-8')
    print(f'Wrote {REPORT}')
    print(f'Current products: {len(current)}; workbook products: {len(product_rows)}; exact title matches: {len(exact_matches)}; sku rows: {sum(sku_counts.values())}; images: {sum(image_counts.values())}; platforms: {sum(platform_counts.values())}')


if __name__ == '__main__':
    main()
