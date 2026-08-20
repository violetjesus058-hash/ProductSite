import json
from pathlib import Path
from openpyxl import load_workbook

ids = {'7542798673','7545217096','7574799746','7543309413'}
root=Path('/home/ubuntu/product-catalog-site')
text=(root/'client/src/data/products.ts').read_text()
start=text.index('export const products: Product[] = ')+len('export const products: Product[] = ')
end=text.index('] as Product[];',start)+1
rows=json.loads(text[start:end])
for row in rows:
    if row.get('sourceProductId') in ids:
        print('FINAL', {k:row.get(k) for k in ['sourceProductId','catalogName','category','subCategory']})
wb=load_workbook('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx',read_only=True,data_only=True)
ws=wb.active
it=ws.iter_rows(values_only=True)
header=list(next(it)); idx={v:i for i,v in enumerate(header)}
seen=set()
for values in it:
    pid=str(values[idx['product_id']]) if values[idx['product_id']] is not None else ''
    if pid in ids and pid not in seen:
        seen.add(pid)
        print('SOURCE', {key: values[idx[key]] if idx.get(key) is not None else None for key in ['product_id','title_original','title_en_platform','category','subcategory','variant_properties_raw']})
