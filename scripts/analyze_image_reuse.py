from __future__ import annotations

import json
import re
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook

WORKBOOK = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
DOWNLOAD_REPORT = Path('/home/ubuntu/product-catalog-site/scripts/product-image-download-report.json')
OUTPUT = Path('/home/ubuntu/product-catalog-site/scripts/image-reuse-analysis.md')


def canonical(url: object) -> str:
    if not url:
        return ''
    text = str(url).strip()
    parts = urlsplit(text)
    path = re.sub(r'\?x-oss-process=.*$', '', parts.path)
    return urlunsplit((parts.scheme.lower(), parts.netloc.lower(), path, '', ''))


def main() -> None:
    download = json.loads(DOWNLOAD_REPORT.read_text(encoding='utf-8'))
    existing = {}
    for item in download.get('results', []):
        if item.get('ok') and item.get('url') and item.get('path'):
            existing[canonical(item['url'])] = item['path']

    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    image_urls = set()
    product_image_counts = {}
    for ws in wb.worksheets:
        if ws.title != 'product_images':
            continue
        rows = ws.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration:
            continue
        headers = [str(v).strip() if v is not None else '' for v in header_values]
        index = {name: i for i, name in enumerate(headers)}
        for values in rows:
            url = values[index['image_url']] if 'image_url' in index and index['image_url'] < len(values) else None
            pid = values[index['product_id']] if 'product_id' in index and index['product_id'] < len(values) else None
            normalized = canonical(url)
            if normalized:
                image_urls.add(normalized)
                if pid is not None:
                    product_image_counts[str(pid)] = product_image_counts.get(str(pid), 0) + 1

    exact_reuse = image_urls & set(existing)
    new_downloads = image_urls - set(existing)
    report = [
        '# 图片增量复用分析', '',
        '| 指标 | 数量 | 说明 |', '|---|---:|---|',
        f'| 当前已成功托管的源图片映射 | {len(existing)} | 来自现有 276 个商品的下载报告 |',
        f'| 新工作簿图片记录 | {sum(product_image_counts.values())} | `product_images` 工作表行数 |',
        f'| 新工作簿去重后的图片 URL | {len(image_urls)} | 按规范化 URL 去重 |',
        f'| 可直接复用的现有图片 URL | {len(exact_reuse)} | 无需重新下载，直接使用已有托管路径 |',
        f'| 需要进一步处理的图片 URL | {len(new_downloads)} | 只有这些才进入增量下载/指纹比对队列 |',
        f'| URL 级直接复用率 | {len(exact_reuse)/max(1,len(image_urls)):.1%} | 仅按 URL 计算，不包含视觉相似图 |',
        '',
        '## 建议', '',
        '优先使用 `product_images.image_url` 与现有下载报告的规范化 URL 做精确映射。精确命中的图片直接复用现有 `/product-images/...webp` 托管路径；只有未命中的 URL 才进行增量下载。对未命中图片，可先下载主图并计算 SHA-256/感知哈希，再决定是否需要详情图。', '',
        '新表的 `sku_records` 含有大量重复商品级图片信息，不应按 SKU 行下载图片；图片应只从 `product_images` 表读取，并以 `product_id + image_order` 聚合。', '',
    ]
    OUTPUT.write_text('\n'.join(report) + '\n', encoding='utf-8')
    print(f'Wrote {OUTPUT}')
    print(f'Existing mappings: {len(existing)}; workbook unique image URLs: {len(image_urls)}; exact reuse: {len(exact_reuse)}; new candidates: {len(new_downloads)}')


if __name__ == '__main__':
    main()
