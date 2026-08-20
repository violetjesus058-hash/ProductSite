import json
from collections import Counter
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path('/home/ubuntu/product-catalog-site')
source_text = (ROOT / 'client/src/data/products.ts').read_text(encoding='utf-8')
start = source_text.index('export const products: Product[] = ') + len('export const products: Product[] = ')
end = source_text.index('] as Product[];', start) + 1
products = json.loads(source_text[start:end])
by_id = {str(p.get('sourceProductId')): p for p in products}

path = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb.active
rows = ws.iter_rows(values_only=True)
headers = list(next(rows))
idx = {name: i for i, name in enumerate(headers)}
matched = []
for row in rows:
    pid = str(row[idx['product_id']]) if row[idx['product_id']] is not None else ''
    if pid in by_id:
        matched.append((by_id[pid], {key: row[value] for key, value in idx.items()}))

by_product = {}
for product, row in matched:
    by_product.setdefault(str(product['sourceProductId']), (product, row))
print('products', len(products), 'source sku rows', len(matched), 'source products matched', len(by_product))
unique_matched = list(by_product.values())
for p, row in unique_matched[:80]:
    print(json.dumps({'id':p['sourceProductId'],'current':p.get('catalogName'),'original':row.get('title_original'),'platform':row.get('title_en_platform'),'subcategory':row.get('subcategory'),'variant':row.get('variant_properties_raw')}, ensure_ascii=False))
print('original title availability', sum(bool(row.get('title_original')) for _,row in unique_matched))
print('platform title availability', sum(bool(row.get('title_en_platform')) for _,row in unique_matched))
print('current generic', sum((p.get('catalogName') or '') in {'Essential Apparel','Everyday Apparel'} for p,_ in unique_matched))
