from __future__ import annotations

import json
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageFilter, ImageStat

ROOT = Path('/home/ubuntu/product-catalog-site')
WORKBOOK = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
DOWNLOAD_REPORT = ROOT / 'scripts/product-image-download-report.json'
COVER_LIST = ROOT / 'scripts/kakobuy-cover-download-list.json'
OUTPUT_JSON = ROOT / 'scripts/kakobuy-full-products.json'
# products.ts is the active frontend data module imported by Home.tsx and detail pages.
OUTPUT_TS = ROOT / 'client/src/data/products.ts'


def money(value: object) -> str:
    try:
        return f'{Decimal(str(value)).quantize(Decimal("0.01"))}'
    except (InvalidOperation, TypeError, ValueError):
        return ''


def canon(url: object) -> str:
    return re.sub(r'\?.*$', '', str(url or '').strip()).lower()


def platform_links_for(source_id: str) -> dict[str, str]:
    """Build additional agent-platform URLs from the existing Weidian ID only."""
    weidian_url = f'https://weidian.com/item.html?itemID={source_id}'
    return {
        'Litbuy': f'https://litbuy.com/product/weidian/{source_id}?inviteCode=XXGYH4Z80',
        'GTbuy': f'https://gtbuy.com/product/weidian/{source_id}?inviteCode=XO78PVRZW',
        'Oopbuy': f'https://oopbuy.com/product/weidian/{source_id}?inviteCode=Y5DH4UF2W',
        'Hipobuy': f'https://hipobuy.com/product/weidian/{source_id}?inviteCode=P6PP29VX7',
        'Fansbuy': f'https://fansbuy.com/item-micro-{source_id}.html?promotionCode=R0dfTU9DRzA2VTk',
        'LoveGoBuy': f'https://www.lovegobuy.com/product?shop_type=weidian&id={source_id}&invite_code=U577HX',
        'Hoobuy': f'https://hoobuy.com/product/2/{source_id}?inviteCode=K8l2grxX',
        'UsFans': f'https://www.usfans.com/product/3/{source_id}?ref=BCSLQC',
        'AllChinaBuy': f'https://www.allchinabuy.com/en/page/buy/?nTag=Home-search&from=search-input&_search=url&position=&url={weidian_url}&partnercode=EEa5go',
        'Mulebuy': f'https://mulebuy.com/product/?shop_type=weidian&id={source_id}&ref=200209428',
        'AcBuy': f'https://www.acbuy.com/product?id={source_id}&source=WD&u=3FLN4S',
        'Joyagoo': f'https://joyagoo.com/product?platform=WEIDIAN&id={source_id}&ref=300950678',
        'OrientDig': f'https://orientdig.com/product/?shop_type=weidian&id={source_id}&ref=100245718',
        'Sugargoo': f'https://www.sugargoo.com/products?productLink={weidian_url}&memberId=3229305473717352480',
        'BBDBuyEU': f'https://www.bbdbuyeu.com/goods/WEIDIAN/{source_id}?inviteCode=j4zwj7',
        'VigorBuy': f'https://vigorbuy.com/product/2/{source_id}?inviteCode=jkNlpqAP',
        'Fishgoo': f'https://www.fishgoo.com/#/product?productLink={weidian_url}&memberId=TG2665ux5KieI',
    }


_IMAGE_QUALITY_CACHE: dict[str, float] = {}


def image_quality_score(image_path: str) -> float:
    """Score cover suitability: blank/near-uniform images are rejected, clear single images score lower."""
    if image_path in _IMAGE_QUALITY_CACHE:
        return _IMAGE_QUALITY_CACHE[image_path]
    try:
        path = ROOT / 'client/public' / image_path.lstrip('/')
        with Image.open(path) as source:
            image = source.convert('L').resize((48, 48))
        pixels = list(image.getdata())
        stats = ImageStat.Stat(image)
        mean = stats.mean[0]
        deviation = stats.stddev[0]
        near_white = sum(pixel >= 248 for pixel in pixels) / max(1, len(pixels))
        near_black = sum(pixel <= 7 for pixel in pixels) / max(1, len(pixels))
        # Pure white/black placeholders and blank panels must never remain covers.
        if deviation < 4.5 or near_white > 0.985 or near_black > 0.985:
            score = 10000.0
        else:
            score = collage_score(image_path)
        _IMAGE_QUALITY_CACHE[image_path] = score
        return score
    except Exception:
        _IMAGE_QUALITY_CACHE[image_path] = 10000.0
        return 10000.0


def preferred_single_image(images: list[str]) -> int | None:
    """Choose a clear later gallery image when the first image is blank or materially worse."""
    if len(images) < 2:
        return None
    candidates = list(enumerate(images[:16]))
    scores = [(image_quality_score(image), index) for index, image in candidates]
    best_score, best_index = min(scores)
    current_score = image_quality_score(images[0])
    if best_index != 0 and (current_score >= 10000.0 or best_score + 3.0 < current_score):
        return best_index
    return None


def clean_title(value: str, fallback: str) -> str:
    value = re.sub(r'📏.*$', '', value)
    value = re.sub(r'whatsapp[:：]?\s*\+?[\d\s().+\-（）]+', '', value, flags=re.I)
    value = re.sub(r'(?:pls\s+add(?:\s+whatsapp)?|pls\s+contact|please\s+contact|if\s+(?:u|you)\s+need(?:\s+(?:size\s+)?recommendation)?).*$', '', value, flags=re.I)
    value = re.sub(r'\b\d{7,}\b', '', value)
    value = re.sub(r'[（(]\s*[）)]', '', value)
    value = re.sub(r'[（(]\s*$', '', value)
    value = re.sub(r'\s+', ' ', value).strip(' -,:;，；')
    return value or fallback


def editorial_descriptor(category: str, subcategory: str, title: str) -> str:
    """Add a restrained catalog-style use-case phrase only when the source title is generic."""
    text = f'{title} {subcategory}'.lower()
    if category == 'pants' and re.search(r'\bshorts\b', text): return 'Warm-Weather'
    if category == 'clothing' and re.search(r'\b(?:shirt|tee|t-?shirt|jersey)\b', text): return 'Everyday'
    if category == 'clothing' and re.search(r'\b(?:hoodie|jacket|sweater|coat)\b', text): return 'Layering'
    if category == 'shoe' and re.search(r'\b(?:sneaker|shoe|boot|sandal)\b', text): return 'Everyday'
    if category == 'pants': return 'Everyday'
    if category in {'bags', 'ACC'} and re.search(r'\b(?:wallet|belt|card holder|bag|bracelet|watch|glasses)\b', text): return 'Daily Essential'
    if category == 'bags': return 'Daily Carry'
    if category == 'ACC': return 'Daily Essential'
    if category == 'fragrance': return 'Everyday'
    if category == 'watches': return 'Everyday'
    if category == 'clothing': return 'Everyday'
    if category == 'pants': return 'Everyday'
    if category == 'shoe': return 'Everyday'
    return ''


def finalize_display_title(title: str, category: str, subcategory: str) -> str:
    """Keep card titles concise while preserving source wording and factual category terms."""
    title = re.sub(r'^\*?\s*\d{2}-\d{2}(?:[A-Z]{1,5})?\s*(?:[/|:-]\s*)?', '', title, flags=re.I)
    title = re.sub(r'\s+[A-Z]{1,5}[-_]\d{2,}\s*$', '', title, flags=re.I)
    title = re.sub(r'^[A-Z]\s+(?=(?:Fashion|Classic|Premium|Versatile)\b)', '', title, flags=re.I)
    title = re.sub(r'\s+', ' ', title).strip(' -/:')
    generic = bool(re.match(r'^(?:high[- ]?quality|rep high[- ]?quality|fashion|premium|versatile|classic|catalog item|kakobuy product)\b', title, re.I))
    descriptor = editorial_descriptor(category, subcategory, title)
    if generic and descriptor:
        title = re.sub(r'^(?:high[- ]?quality|rep high[- ]?quality|fashion|premium|versatile|classic)\s*', '', title, flags=re.I).strip(' -')
        if not title or re.fullmatch(r'(?:fashion|versatile|apparel|product|item)', title, re.I):
            category_label = {'clothing': 'Apparel', 'pants': 'Pants', 'shoe': 'Shoes', 'bags': 'Bag', 'fragrance': 'Perfume', 'watches': 'Watch', 'ACC': 'Accessory'}.get(category, 'Product')
            title = subcategory if subcategory and subcategory.lower() != 'selection' else category_label
        title = f'{descriptor} {title}'.strip()
    if len(title) <= 42:
        return title
    shortened = re.sub(r'\s*[-|/:].*$', '', title).strip()
    if len(shortened) <= 42:
        return shortened
    words = shortened.split()
    compact = ''
    for word in words:
        candidate = f'{compact} {word}'.strip()
        if len(candidate) > 39: break
        compact = candidate
    return compact or title[:39].rstrip()


def conservative_category_title(category: str, subcategory: str, fallback: str) -> str:
    sub = str(subcategory or '').strip().lower()
    if category == 'pants':
        return {'trousers': 'Everyday Trousers', 'jeans': 'Everyday Jeans', 'sweatpants': 'Everyday Sweatpants', 'shorts': 'Everyday Shorts'}.get(sub, 'Everyday Pants')
    if category == 'clothing':
        return {'jackets': 'Everyday Jacket', 'shirts': 'Everyday Shirt', 'hoodies': 'Everyday Hoodie', 'sweaters': 'Everyday Sweater', 'sets': 'Everyday Clothing Set'}.get(sub, 'Everyday Apparel')
    if category == 'shoe':
        return {'sneakers': 'Everyday Sneakers', 'sandals': 'Everyday Sandals', 'boots': 'Everyday Boots'}.get(sub, 'Everyday Shoes')
    if category == 'ACC':
        return 'Everyday Accessories'
    if category == 'bags':
        return 'Everyday Bag'
    if category == 'fragrance':
        return 'Everyday Fragrance'
    if category == 'watches':
        return 'Everyday Watch'
    return fallback


def optimize_display_title(info: dict, category: str, subcategory: str, fallback: str) -> str:
    """Use source wording when meaningful; replace ID-like titles with concise evidence-based names."""
    title = clean_title(str(info.get('title', '') or ''), '')
    original = clean_title(str(info.get('title_original', '') or ''), '')
    source_text = f'{title} {original}'.lower()
    size_only_source = bool(re.fullmatch(r'[*]?\s*\d+\s*[-–]\s*\d+(?:\s+\d+)?(?:\s+[A-Z]{1,5})?', title, re.I))
    if size_only_source:
        sub_label = {'Underwear': 'Underwear', 'Sets': 'Clothing Set', 'Hoodies': 'Hoodie', 'Jackets': 'Jacket', 'Sweaters': 'Sweater', 'Shirts': 'Shirt', 'Caps': 'Cap', 'Watches': 'Watch', 'Sandals': 'Sandals'}.get(subcategory)
        category_label = {'clothing': 'Apparel', 'pants': 'Pants', 'shoe': 'Shoes', 'bags': 'Bag', 'fragrance': 'Perfume', 'watches': 'Watch', 'ACC': 'Accessory'}.get(category, 'Product')
        return finalize_display_title(f'High-Quality {sub_label or category_label}', category, subcategory)
    if re.search(r'[\u4e00-\u9fff]', title) or re.search(r'[\u4e00-\u9fff]', original):
        translated = original or title
        translations = [
            ('高品质', 'High-Quality'), ('高质量', 'High-Quality'), ('精品', 'Premium'), ('时尚', 'Fashion'), ('经典', 'Classic'), ('百搭', 'Versatile'),
            ('皮带', 'Belt'), ('腰带', 'Belt'), ('包包', 'Bag'), ('内裤', 'Underwear'), ('套装', 'Clothing Set'), ('卫衣', 'Hoodie'), ('衬衫', 'Shirt'),
            ('短袖', 'Short-Sleeve Shirt'), ('长袖', 'Long-Sleeve Shirt'), ('牛仔裤', 'Jeans'), ('裤子', 'Pants'), ('裤', 'Pants'), ('帽子', 'Hat'), ('帽', 'Hat'),
            ('运动鞋', 'Sneakers'), ('鞋子', 'Shoes'), ('鞋', 'Shoes'), ('手表', 'Watch'), ('香水', 'Perfume'), ('手机壳', 'Phone Case'), ('眼镜', 'Glasses'),
            ('手链', 'Bracelet'), ('项链', 'Necklace'), ('袜子', 'Socks'), ('袜', 'Socks'), ('羽绒服', 'Puffer Jacket'), ('夹克', 'Jacket'), ('外套', 'Jacket'),
            ('毛衣', 'Sweater'), ('围巾', 'Scarf'), ('钱包', 'Wallet'), ('卡包', 'Card Holder'), ('耳机', 'Headphones'), ('裙子', 'Skirt'), ('裙', 'Skirt'), ('背心', 'Vest'),
        ]
        for token, label in translations:
            translated = translated.replace(token, f' {label} ')
        translated = re.sub(r'[\u4e00-\u9fff]', ' ', translated)
        translated = re.sub(r'[_:：；;]+', ' ', translated)
        translated = re.sub(r'[（(（]\s*(?:[A-Z]{1,5}[-_]\d+|\d+[-_][A-Z]{1,5}[-_]\d+)\s*[）)）]', ' ', translated, flags=re.I)
        translated = re.sub(r'\s+(?:\d+[-_][A-Z]{1,5}[-_]\d+|[A-Z]{1,5}[-_]\d{2,})\s*$', '', translated, flags=re.I)
        if re.match(r'^(?:High-Quality|High Quality|Fashion|Premium)', translated, re.I):
            translated = re.sub(r'\s+\d{1,4}\s*$', '', translated)
        translated = re.sub(r'\b([A-Za-z]+(?:-[A-Za-z]+)?)\s+\1\b', r'\1', translated, flags=re.I)
        translated = re.sub(r'^\*?\s*\d{2}-\d{2}(?:[A-Z]{1,5})?\s*(?:[/|:-]\s*)?', '', translated, flags=re.I)
        translated = re.sub(r'\s+', ' ', translated).strip(' -/:')
        if translated and not re.search(r'[\u4e00-\u9fff]', translated):
            translated_size_only = bool(re.fullmatch(r'[*]?\s*\d+\s*[-–]\s*\d+(?:\s+\d+)?(?:\s+[A-Z]{1,5})?', translated, re.I))
            if translated_size_only:
                sub_label = {'Underwear': 'Underwear', 'Sets': 'Clothing Set', 'Hoodies': 'Hoodie', 'Jackets': 'Jacket', 'Sweaters': 'Sweater', 'Shirts': 'Shirt', 'Caps': 'Cap', 'Watches': 'Watch', 'Sandals': 'Sandals'}.get(subcategory)
                category_label = {'clothing': 'Apparel', 'pants': 'Pants', 'shoe': 'Shoes', 'bags': 'Bag', 'fragrance': 'Perfume', 'watches': 'Watch', 'ACC': 'Accessory'}.get(category, 'Product')
                return finalize_display_title(f'High-Quality {sub_label or category_label}', category, subcategory)
            translated_model_category_like = bool(re.fullmatch(r'(?:[A-Z]{1,8}\d{1,6}|\d{1,8}[A-Z]{1,8}\d*|[A-Z0-9]{2,}[-_.]\d+)(?:\s+|[-_])(?:Pants|Jeans|Shorts|Trousers|Hoodie|Jersey|Shirt|Sweater|Jacket|Shoes?|Sneakers?|Boots?|Bag|Wallet|Belt|Cap|Hat|Socks?|Underwear|Watch|Perfume|Accessory)(?:\s+(?:S-XL|S-L|M-XXL|\d+[-–]\d+))?', translated, re.I))
            if translated_model_category_like:
                fallback_label = conservative_category_title(category, subcategory, fallback)
                return finalize_display_title(fallback_label, category, subcategory)
            return finalize_display_title(translated, category, subcategory)
    generic_title = re.match(r'^(?:catalog item|kakobuy product|all[-_ ]?[a-z]+|high[- ]?quality|high-quality|rep high[- ]?quality|rep high quality|fashion hat|fashion trend|high-quality fashion)', title, re.I)
    if generic_title:
        title = re.sub(r'\s*(?:[（(]\s*)?(?:\d+[-_][A-Z]{1,5}[-_]\d+|[A-Z]{1,5}[-_]\d{2,}|\d{1,4})\s*(?:[）)])?\s*$', '', title, flags=re.I).strip(' -')
    size_only = bool(re.fullmatch(r'[*]?\s*\d+\s*[-–]\s*\d+(?:\s+\d+)?(?:\s+[A-Z]{1,5})?', title, re.I))
    generic_exact = size_only or bool(re.fullmatch(r'(?:all|rep high[- ]?quality|high[- ]?quality|gs\.\d+|[a-z]{1,5}\.[0-9]{4,})', title, re.I))
    identifier_like = generic_exact or bool(re.fullmatch(r'[A-Z0-9]{2,}[._-]?', title, re.I)) or bool(re.search(r'\b(?:catalog item|kakobuy product)\b|(?:all[-_ ]?[a-z]+|high[- ]?quality|high-quality|rep high[- ]?quality|rep high quality)\s*\d*[-_ ]?(?:[a-z]{1,5}[-_ ]?\d+)?|\b\d+[-_][a-z]{1,5}[-_ ]?\d+\b', title, re.I))
    model_category_like = bool(re.fullmatch(r'(?:[A-Z]{1,8}\d{1,6}|\d{1,8}[A-Z]{1,8}\d*|[A-Z0-9]{2,}[-_.]\d+)(?:\s+|[-_])(?:Pants|Jeans|Shorts|Trousers|Hoodie|Jersey|Shirt|Sweater|Jacket|Shoes?|Sneakers?|Boots?|Bag|Wallet|Belt|Cap|Hat|Socks?|Underwear|Watch|Perfume|Accessory)(?:\s+(?:S-XL|S-L|M-XXL|\d+[-–]\d+))?', title, re.I))
    if model_category_like:
        fallback_label = conservative_category_title(category, subcategory, fallback)
        return finalize_display_title(fallback_label, category, subcategory)
    if not identifier_like:
        return finalize_display_title(title or fallback, category, subcategory)

    code_match = re.search(r'\b(?:\d+[-_])?([A-Z]{1,5})[-_]\d+\b', title, re.I)
    code = code_match.group(1).upper() if code_match else ''
    code_labels = {
        'TS': 'T-Shirt', 'TST': 'T-Shirt', 'ST': 'T-Shirt', 'SS': 'Short-Sleeve Shirt',
        'HD': 'Hoodie', 'HDS': 'Hoodie', 'SW': 'Sweater', 'JE': 'Jeans', 'LO': 'Lounge Pants',
        'PT': 'Pants', 'HB': 'Handbag', 'FH': 'Fashion Hat', 'KH': 'Knit Hat', 'UN': 'Underwear',
        'PHD': 'Puffer Jacket', 'VT': 'Vest', 'RC': 'Jacket', 'CL': 'Clothing Set',
    }
    chinese_labels = [('内裤', 'Underwear'), ('卫衣', 'Hoodie'), ('套装', 'Clothing Set'), ('牛仔裤', 'Jeans'), ('裤', 'Pants'), ('帽', 'Hat'), ('鞋', 'Shoes'), ('包', 'Bag'), ('手表', 'Watch'), ('香水', 'Perfume')]
    for token, label in chinese_labels:
        if token in original:
            return finalize_display_title(f'High-Quality {label}', category, subcategory)
    if code in code_labels:
        return finalize_display_title(f'High-Quality {code_labels[code]}', category, subcategory)
    category_labels = {'clothing': 'Apparel', 'pants': 'Pants', 'shoe': 'Shoes', 'bags': 'Bag', 'fragrance': 'Perfume', 'watches': 'Watch', 'ACC': 'Accessory'}
    sub_label = {'Underwear': 'Underwear', 'Sets': 'Clothing Set', 'Hoodies': 'Hoodie', 'Jackets': 'Jacket', 'Sweaters': 'Sweater', 'Shirts': 'Shirt', 'Caps': 'Cap', 'Watches': 'Watch', 'Sandals': 'Sandals'}.get(subcategory)
    return finalize_display_title(f'High-Quality {sub_label or category_labels.get(category, "Product")}', category, subcategory)


CATEGORY_PATTERNS = {
    'shoe': r'\b(shoes?|sneakers?|boots?|slides?|sandals?|loafers?|mules?|running shoes?)\b',
    'watches': r'\b(watch(?:es)?|smart watch(?:es)?|mechanical watch(?:es)?|腕表|手表)\b',
    'fragrance': r'\b(perfumes?|fragrances?|colognes?|parfum|香水)\b',
    'bags': r'\b(bags?|backpacks?|totes?|shoulder bags?|crossbody|handbags?|pouches?)\b',
    'wallets': r'\b(wallets?|card holders?|cardholder|keychains?|零钱包|钱包)\b',
    'pants': r'\b(pants|trousers|jeans|shorts|joggers|sweatpants|leggings|denim|cargo|bottoms|长裤|短裤|牛仔裤|休闲裤|运动裤)\b',
    'clothing': r'\b(shirts?|t-?shirts?|tees?|short sleeve|short-sleeve|hoodies?|jackets?|coats?|sweaters?|sweatshirts?|jerseys?|sportswear|tracksuits?|dresses?|skirts?|vests?|tops?|clothing|衬衫|卫衣|夹克|外套|毛衣|短袖|长袖|裙)\b',
    'accessories': r'\b(caps?|hats?|belts?|glasses|sunglasses|scarves?|ties?|bracelets?|rings?|necklaces?|earrings?|headphones?|phone cases?|accessor)\b',
}
BRAND_PATTERNS = [('Louis Vuitton', r'\b(louis\s*vuitton|lv)\b'), ('Stone Island', r'\bstone\s*island\b'), ('Ralph Lauren', r'\bralph\s*lauren\b'), ('Nike', r'\bnike\b'), ('Adidas', r'\badidas\b'), ('Puma', r'\bpuma\b'), ('Balenciaga', r'\bbalenciaga\b'), ('Dior', r'\bdior\b'), ('Moncler', r'\bmoncler\b'), ('The North Face', r'\bthe\s*north\s*face\b'), ('New Balance', r'\bnew\s*balance\b'), ('Gucci', r'\bgucci\b'), ('Prada', r'\bprada\b'), ('Supreme', r'\bsupreme\b')]
# Confirmed visual findings only. Suspected cases remain separate and never alter category fields.
SUSPECTED_REVIEW = {
    # Add sourceProductId: note when a product looks suspicious but the cover alone is not conclusive.
    '7543327097': {'review_note': 'Pants audit: thumbnail/title indicate a shorts-and-short-sleeve mixed set; retain current Pants/Shorts classification pending detail-page confirmation.'},
    '7773126421': {'review_note': 'Pants audit: long-sleeve-and-long-pants set is mixed apparel; retain current Pants/Trousers classification pending detail-page confirmation.'},
    '7545310284': {'review_note': 'Pants audit: hooded sweatshirt, sweatpants, and down-jacket set is mixed apparel; retain current Pants/Sweatpants classification pending detail-page confirmation.'},
    '7776165618': {'review_note': 'Pants audit: thumbnail shows a shorts-and-short-sleeve mixed set; retain current Pants/Shorts classification pending detail-page confirmation.'},
    '7545203334': {'review_note': 'Accessories audit: thumbnail shows a mixed sportswear set; retain current Accessories classification pending full detail-page confirmation.'},
}

AI_TITLE_OVERRIDES = {
    '7542974543': 'Urban Streetwear Outfit',
    '7543046447': 'Graphic Print T-Shirts',
    '7543309413': 'Assorted Low-Cut Socks',
    '7543313325': 'Assorted Logo T-Shirts',
    '7543327115': 'Black Textured Metal-Buckle Belt',
    '7543332937': 'Printed Crewneck T-Shirt',
    '7543336857': 'Double GG Buckle Belt',
    '7543342823': 'Black Full-Zip Hoodie',
    '7543346835': 'Cable-Knit Quarter-Zip Sweaters',
    '7543348805': 'Compass Graphic T-Shirt',
    '7543364725': 'White Polo with Red-Navy Trim',
    '7543368735': 'Black Logo Crewneck Sweatshirt',
    '7543368745': 'Heart Embroidered Hoodie Set',
    '7543374625': 'Black Crewneck Sweater Red Heart',
    '7543382897': 'Graphic T-Shirt Selection',
    '7543386603': 'Black Hooded Jacket',
    '7543398543': 'Everyday Lightweight No-Show Socks',
    '7544656828': 'Hoodie and Sweatpants Set',
    '7545207126': 'Blue and Black Hoodie',
    '7545217100': 'Light Gray Crewneck Sweater',
    '7545221002': 'Everyday Apparel',
    '7545221022': 'Graphic Logo T-Shirts',
    '7545238802': 'Letter Logo T-Shirt',
    '7545260684': 'Everyday Hoodie',
    '7545270868': 'Grey Quilted Snap-Front Vest',
    '7545272892': 'Logo Print T-Shirts',
    '7545294490': 'Quarter-Zip Pullover Sweaters',
    '7545296400': 'Gray Jogger Sweatpants',
    '7545312262': 'Gray Zip Hoodie & Joggers',
    '7545316058': 'Floral Hoodie & Jogger Set',
    '7547224894': 'Everyday Fragrance',
    '7552721905': 'Monogram Jacquard Belt',
    '7554525338': 'Everyday Cap',
    '7574787017': 'Ribbed Logo Tank Tops',
    '7574802935': 'Quarter-Zip Athletic Tops',
    '7574808617': 'Graphic Tee Selection',
    '7576483261': 'Assorted Patterned Bags',
    '7576491587': 'Daily Essential Belt',
    '7576538857': 'Ribbed Tank Tops',
    '7576546619': 'Black Croc-Embossed Belt',
    '7576548605': 'Black Eau de Parfum',
    '7576550609': 'Floral Print Hoodie Set',
    '7576550615': 'Everyday Sneakers',
    '7576552591': 'Blue Hoodie and Sweatpants Set',
    '7576554555': 'Ripped Light Wash Jeans',
    '7576564461': 'Clothing Size Chart (S–XL)',
    '7576566371': 'Who Cares Novelty Watch',
    '7576570407': 'Graphic T‑Shirt Selection',
    '7576574207': 'Graphic Hoodies & Joggers',
    '7576574219': 'Grey Snap-Front Puffer Vest',
    '7576584181': 'Black Graphic Hoodie',
    '7576593843': 'Beige Plaid Fringe Scarf',
    '7576597943': 'Monogram Shield Sunglasses',
    '7576599901': 'Crocodile Motif Chain Bracelet',
    '7576605827': 'Assorted Track Jackets & Pants',
    '7576605833': 'Portable Speaker',
    '7576607783': 'Assorted Crew Neck T-Shirts',
    '7576609789': 'Grey LA Logo Cap',
    '7576617677': 'H Buckle Belt',
    '7576619633': 'Assorted Eyewear Frames',
    '7576619639': 'Puffer Jackets & Hoodies',
    '7576619721': 'Ape Head Graphic Tee',
    '7576679368': 'Graphic Crewneck Tees',
    '7576686904': 'Assorted Graphic T-Shirts',
    '7576695090': 'Beige Plaid Drawstring Shorts',
    '7576702822': 'City Polo Shirts',
    '7576704688': 'Assorted Graphic T-Shirts',
    '7576742276': 'Soccer Training Tracksuit Sets',
    '7576767828': 'Clothing Selection Image',
    '7576787728': 'Training Tracksuit Sets',
    '7578399472': 'Black Striped Tracksuit Set',
    '7578406954': 'Hoodie and Sweatpants Set',
    '7578419186': 'Track Jacket and Pants',
    '7578446804': 'Clothing Selection - Unclear',
    '7578456650': 'Assorted Graphic T-Shirts',
    '7578458552': 'Black Hooded Jacket Set',
    '7578460522': 'Slim Oval Buckle Belt',
    '7578468476': 'Ornate Link Bracelet',
    '7578468480': 'Puffer Jackets',
    '7578470372': 'Long-Sleeve Training Tracksuit Set',
    '7578482250': 'Smartwatch Selection',
    '7578496024': 'Everyday Fragrance',
    '7578503954': 'Graphic Crewneck T-Shirt',
    '7578505866': 'Monogram Tote & Pouch Set',
    '7578535932': 'Printed Wallets & Card Holders',
    '7601623089': 'Gray & Olive Zip-Up Hoodies',
    '7603539016': 'Imagination Graphic Long Sleeve',
    '7603548814': 'Everyday Jeans',
    '7603560398': 'Pullover Hoodies - Gray Cream Blue',
    '7603586106': 'Fleetwood Mac Graphic Henleys',
    '7611840843': 'Everyday Sneakers',
    '7612177793': 'Daily Essential Belt',
    '7615053508': 'Black Croc-Embossed Belt',
    '7615155794': 'Slim Oval-Buckle Belt',
    '7615163710': 'Grey LA Logo Baseball Cap',
    '7778049033': 'Black Slip-On Clog',
    '7782557155': 'Barcelona Home Fan Version',
    '7782631879': 'Everyday Jersey',
    '7786821585': 'Assorted Silver Rings',
}

MANUAL_OVERRIDES = {
    '7576679472': {'primary_image_index': 3, 'review_note': 'Manual image review: replace size-chart cover with a clear single-polo gallery image.'},
    '7545224884': {'primary_image_index': 2, 'review_note': 'Manual image review: replace multi-shirt collage with a clear single-shirt gallery image.'},
    '7574761499': {'primary_image_index': 0, 'review_note': 'Manual image review: replace mixed cover with a clear single-shirt gallery image.'},
    '7543307459': {'category': 'clothing', 'subCategory': 'Shirts', 'primary_image_index': 2, 'review_note': 'Manual review: use third gallery image as catalog cover.'},
    '7576530885': {'primary_image_index': 2, 'remove_image_indices': [0], 'review_note': 'Manual image review: remove the original first image and promote the original fourth gallery image to the cover.'},
    '7576623430': {'primary_image_index': 6, 'review_note': 'Manual image review: replace blank/size-chart cover with the clear single-shirt gallery image.'},
    '7543348805': {'primary_image_index': 4, 'review_note': 'Manual image review: replace blank cover with a clear single-shirt gallery image.'},
    '7543366669': {'category': 'ACC', 'subCategory': 'Caps', 'primary_image_index': 4, 'review_note': 'Clothing first-image audit: cover clearly shows a knit beanie assortment; corrected to accessories/caps and use a clean single-beanie gallery image.'},
    '7543332937': {'primary_image_index': 4, 'review_note': 'Manual image review: replace blank/size-chart cover with the clear single-shirt gallery image.'},
    '7576584181': {'primary_image_index': 1, 'review_note': 'Manual image review: use the clear single-hoodie gallery image as the catalog cover.'},
    '7576695090': {'primary_image_index': 3, 'review_note': 'Manual image review: use the clear single-shorts gallery image; exclude contact and size-chart images.'},
    '7578440914': {'primary_image_index': 2, 'review_note': 'Manual image review: replace size-chart cover with a clear single-shoe gallery image; account for the downloaded cover entry.'},
    '7611885929': {'primary_image_index': 2, 'review_note': 'Manual image review: replace size-chart cover with a clear single-shoe gallery image; account for the downloaded cover entry.'},
    '7545254650': {'primary_image_index': 2, 'review_note': 'Manual image review: replace size-chart cover with a clear single-shoe gallery image; account for the downloaded cover entry.'},
    '7543424263': {'primary_image_index': 2, 'review_note': 'Shoes image review: replace size-chart cover with a clear single-shoe gallery image; account for the downloaded cover entry.'},
    '7543386595': {'primary_image_index': 1, 'review_note': 'Shoes image review: replace size-chart cover with a clear single-sneaker gallery image.'},
    '7576647651': {'primary_image_index': 2, 'review_note': 'Shoes image review: replace size-chart/color-grid cover with a clear single-shoe gallery image.'},
    '7614865968': {'primary_image_index': 2, 'review_note': 'Shoes image review: replace size-chart/color-grid cover with a clear single-shoe gallery image.'},
    '7576542737': {'primary_image_index': 2, 'review_note': 'Shoes image review: replace product/contact collage with a clear single-boot gallery image.'},
    '7576582183': {'primary_image_index': 1, 'review_note': 'Shoes image review: replace size-chart cover with a clear single-shoe gallery image.'},
    '7578513696': {'primary_image_index': 2, 'review_note': 'Shoes image review: replace watermark/contact-heavy cover with a clean single-shoe gallery image.'},
    '7611891953': {'primary_image_index': 2, 'review_note': 'Shoes image review: replace watermark/contact-heavy cover with a clean single-shoe gallery image.'},
    '7576597949': {'primary_image_index': 2, 'review_note': 'Shoes image review: replace Many Colors collage with a clean single-shoe gallery image.'},
    '7611888011': {'primary_image_index': 2, 'review_note': 'Shoes image review: replace Many Colors collage with a clean single-shoe gallery image.'},
    '7611881149': {'primary_image_index': 2, 'review_note': 'Shoes image review: replace product/contact collage with a clear single-boot gallery image.'},
    '7578496024': {'primary_image_path': '/product-images/7578496024-single-bottle.webp', 'review_note': 'Manual image review: use a source-derived single-bottle crop as the catalog cover; retain the original collage in the gallery.'},
    '7786196426': {'primary_image_path': '/product-images/7786196426-single-watch.webp', 'review_note': 'Manual image review: use a source-derived single-watch crop as the catalog cover; retain the original collage in the gallery.'},
    '7778863173': {'primary_image_path': '/manus-storage/kb-7778863173-cover_a7bcd7b2.gif', 'review_note': 'User-provided cover override: use the supplied GIF as the catalog main image because the generated cover was blank.'},
    '7576540787': {'category': 'pants', 'subCategory': 'Shorts', 'primary_image_index': 2, 'review_note': 'User screenshot review: the first gallery image is blank; use the later single-shorts image as the catalog cover and classify as pants/shorts.'},
    '7576599899': {'category': 'pants', 'subCategory': 'Shorts', 'review_note': 'User screenshot review: title and visible product image identify a pants/shorts item, not a generic clothing item; corrected to pants/shorts.'},
    '7601623089': {'category': 'clothing', 'subCategory': 'Hoodies', 'review_note': 'Manual review: cover image shows two hooded zip-up sweatshirts.'},
    '7603560398': {'category': 'clothing', 'subCategory': 'Hoodies', 'review_note': 'Manual review: cover image shows three hooded sweatshirts.'},
    '7578517746': {'category': 'clothing', 'subCategory': 'Jackets', 'review_note': 'Manual review: cover image shows a padded jacket; retain body-weight size labels.'},
    '7603548814': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: misclassified as clothing, corrected to pants/jeans.'},
    '7578442864': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: misclassified as clothing, corrected to pants/trousers.'},
    '7578482254': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: misclassified as clothing, corrected to pants/jeans.'},
    '7574783515': {'category': 'ACC', 'subCategory': 'Caps', 'review_note': 'AI Audit: misclassified as clothing, corrected to accessories/caps.'},
    '7603616148': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: misclassified as clothing, corrected to pants/trousers.'},
    '7576524975': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: misclassified as clothing, corrected to pants/jeans.'},
    '7576507291': {'category': 'pants', 'subCategory': 'Shorts', 'review_note': 'AI Audit: misclassified as clothing, corrected to pants/shorts.'},
    '7578478202': {'category': 'pants', 'subCategory': 'Shorts', 'review_note': 'AI Audit: misclassified as clothing, corrected to pants/shorts.'},
    '7782578621': {'category': 'ACC', 'subCategory': 'Caps', 'review_note': 'AI Audit: misclassified as clothing, corrected to accessories/caps.'},
    '7601668247': {'category': 'pants', 'subCategory': 'Sweatpants', 'review_note': 'AI Audit: misclassified as clothing, corrected to pants/sweatpants.'},
    '7601753745': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: misclassified as clothing, corrected to pants/trousers.'},
    '7603615668': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: misclassified as clothing, corrected to pants/jeans.'},
    '7578509826': {'category': 'pants', 'subCategory': 'Jeans', 'primary_image_index': 6, 'review_note': 'AI Audit: misclassified as clothing, corrected to pants/jeans; manual image review replaces the blank/size-chart cover with a clear single-shirt gallery image.'},
    '7578515778': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: thumbnail clearly shows a pants-only product assortment; corrected to pants/jeans.'},
    '7545207144': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: High-Quality 3-JE-001 thumbnail clearly shows a pants-only assortment; corrected to pants/jeans.'},
    '7601654481': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7603553408': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: thumbnail clearly shows a pants-only cargo trouser product; corrected to pants/trousers.'},
    '7603641850': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7576695114': {'category': 'pants', 'subCategory': 'Shorts', 'review_note': 'AI Audit: thumbnail clearly shows a shorts-only product assortment; corrected to pants/shorts.'},
    '7601743753': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: thumbnail clearly shows a pants-only denim product; corrected to pants/jeans.'},
    '7543396675': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: REP High Quality 4-PT-001 thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7603667852': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: KJ6300 work-trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7576505455': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: REP High Quality 1-JE-001 thumbnail clearly shows a jeans-only assortment; corrected to pants/jeans.'},
    '7576502955': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: thumbnail clearly shows a jeans-only product; corrected to pants/jeans.'},
    '7543371979': {'category': 'accessories', 'subCategory': 'Caps', 'review_note': 'AI Audit: High Quality Knitted Hat from Factory thumbnail clearly shows knit hats only; corrected to accessories/caps.'},
    '7601613265': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: WK339 jeans thumbnail clearly shows a pants-only denim product; corrected to pants/jeans.'},
    '7543352763': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: High Quality 2-JE-002 thumbnail clearly shows a jeans-only product; corrected to pants/jeans.'},
    '7543334883': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'AI Audit: High-Quality Fashion Premium 1-AM-001 thumbnail clearly shows sneakers; corrected to shoes/sneakers.'},
    '7603630018': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: CK9500 curved-leg trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7543327119': {'category': 'accessories', 'subCategory': 'Scarves', 'review_note': 'AI Audit: High Quality Scarf 001 thumbnail clearly shows scarves only; corrected to accessories/scarves.'},
    '7601743943': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: K5512 straight-leg trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7603683608': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: KJ6267 cargo-trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7601688665': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: CK9556 straight-leg trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7543404477': {'category': 'accessories', 'subCategory': 'Caps', 'review_note': 'AI Audit: High Quality Knitted 3-AT-001 thumbnail clearly shows knit caps only; corrected to accessories/caps.'},
    '7603633944': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: CK6606 curved-leg trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7742257053': {'category': 'pants', 'subCategory': 'Shorts', 'review_note': 'AI Audit: thumbnail clearly shows a shorts-only product; corrected to pants/shorts.'},
    '7601666909': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: CK3367 camouflage-trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7603563092': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: CK9621 curved-leg trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7601688735': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: K6110 cargo-trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7603591798': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'AI Audit: NZ7591 denim-trouser thumbnail clearly shows a pants-only product; corrected to pants/jeans.'},
    '7601761663': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: CK9551 straight-leg trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7545274624': {'category': 'accessories', 'subCategory': 'Caps', 'review_note': 'AI Audit: REP High Quality 1-FH-001 thumbnail clearly shows knit caps only; corrected to accessories/caps.'},
    '7601653203': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: K5308 flared-trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7601720055': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: CK6669 curved-leg trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7601734009': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'AI Audit: K5305 flared-trouser thumbnail clearly shows a pants-only product; corrected to pants/trousers.'},
    '7578509830': {'category': 'clothing', 'subCategory': 'Jackets', 'review_note': 'Pants audit: thumbnail clearly shows denim jackets/coats rather than pants; corrected to clothing/jackets.'},
    '7543323077': {'category': 'clothing', 'subCategory': 'Jackets', 'review_note': 'Pants audit: High-Quality Denim Jacket - 001-CN thumbnail clearly shows denim jackets/coats rather than pants; corrected to clothing/jackets.'},
    '7576502949': {'category': 'pants', 'subCategory': 'Shorts', 'review_note': 'Pants audit: thumbnail shows shorts-only assortments despite the generic pants title; corrected subcategory to pants/shorts.'},
    '7785584510': {'category': 'clothing', 'subCategory': 'Shirts', 'review_note': 'Fragrance audit: thumbnail clearly shows a football jersey rather than fragrance; corrected to clothing/shirts.'},
    '7578486178': {'category': 'bags', 'subCategory': 'Shoulder Bags', 'review_note': 'Accessories audit: source title and thumbnail clearly show a shoulder/crossbody bag; corrected to bags/shoulder bags.'},
    '7576570355': {'category': 'bags', 'subCategory': 'Shoulder Bags', 'review_note': 'Accessories audit: source title and thumbnail clearly show a travel handbag; corrected to bags/shoulder bags.'},
    '7578468490': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories audit: title and thumbnail clearly show men’s underwear/briefs; corrected to clothing/underwear.'},
    '7574783515': {'category': 'clothing', 'subCategory': 'Sets', 'review_note': 'Accessories audit: title and thumbnail clearly show a training short-sleeve clothing set; corrected to clothing/sets.'},
    '7578480194': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories audit: title and thumbnail clearly show leather sports shoes; corrected to shoes/sneakers.'},
    '7578480258': {'category': 'fragrance', 'subCategory': 'Perfume', 'review_note': 'Accessories audit: title and thumbnail clearly show perfume; corrected to fragrance/perfume.'},
    '7576493523': {'category': 'watches', 'subCategory': 'Watches', 'review_note': 'Accessories audit: title and thumbnail clearly show a wristwatch; corrected to watches.'},
    '7576582179': {'category': 'bags', 'subCategory': 'Backpacks', 'review_note': 'Accessories audit: title and thumbnail clearly show a backpack; corrected to bags/backpacks.'},
    '7576558465': {'category': 'ACC', 'subCategory': 'Caps', 'review_note': 'Accessories audit: title and thumbnail clearly show a fashion cap; corrected to accessories/caps.'},
    '7576593843': {'category': 'ACC', 'subCategory': 'Scarves', 'review_note': 'Accessories audit: title and thumbnail clearly show a scarf; corrected to accessories/scarves.'},
    '7601699767': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'Clothing audit: title and thumbnail clearly show NZ7346 denim jeans rather than clothing; corrected to pants/jeans.'},
    '7543362649': {'category': 'clothing', 'subCategory': 'Sportswear', 'review_note': 'Accessories audit: title and thumbnail clearly show a training-clothes apparel product; corrected to clothing/sportswear.'},
    '7783164819': {'category': 'clothing', 'subCategory': 'Shirts', 'review_note': 'Accessories audit: title and thumbnail clearly show a long-sleeve apparel product; corrected to clothing/shirts.'},
    '7789908090': {'category': 'clothing', 'subCategory': 'Sportswear', 'review_note': 'Accessories audit: title and thumbnail clearly show fashion apparel rather than an accessory; corrected to clothing/sportswear.'},
    '7782578621': {'category': 'clothing', 'subCategory': 'Shirts', 'review_note': 'Accessories audit: title and thumbnail clearly show an Algeria football jersey; corrected to clothing/shirts.'},
    '7552674207': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories audit: title and thumbnail clearly show underwear; corrected to clothing/underwear.'},
    '7543334869': {'category': 'clothing', 'subCategory': 'Sportswear', 'review_note': 'Accessories audit: title and thumbnail clearly show soccer training apparel; corrected to clothing/sportswear.'},
    '7615081442': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories audit: title and thumbnail clearly show cotton underwear; corrected to clothing/underwear.'},
    '7612091525': {'category': 'clothing', 'subCategory': 'Socks', 'review_note': 'Accessories audit: title and thumbnail clearly show cotton basketball socks; corrected to clothing/socks.'},
    '7578512126': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories audit: title and thumbnail clearly show underwear; corrected to clothing/underwear.'},
    '7614875914': {'category': 'shoe', 'subCategory': 'Sandals', 'review_note': 'Accessories audit: title and thumbnail clearly show casual slippers; corrected to shoes/sandals.'},
    '7543321151': {'category': 'clothing', 'subCategory': 'Sportswear', 'review_note': 'Accessories audit: title and thumbnail clearly show a PSG training suit; corrected to clothing/sportswear.'},
    '7545238798': {'category': 'ACC', 'subCategory': 'Caps', 'review_note': 'Clothing audit: title and thumbnail clearly show a casual cap assortment; corrected to accessories/caps.'},
    '7545272902': {'category': 'ACC', 'subCategory': 'Caps', 'review_note': 'Clothing audit: title and thumbnail clearly show a casual cap assortment; corrected to accessories/caps.'},
    '7545203334': {'category': 'clothing', 'subCategory': 'Sportswear', 'review_note': 'Accessories second-pass audit: image clearly shows long-sleeve athletic tops and matching track pants; corrected to clothing/sportswear.'},
    '7762682043': {'category': 'clothing', 'subCategory': 'Sets', 'review_note': 'Accessories second-pass audit: title and thumbnail clearly show a coordinated apparel outfit; corrected to clothing/sets.'},
    '7554611924': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories second-pass audit: image clearly shows men’s boxer briefs; corrected to clothing/underwear.'},
    '7578488198': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'Accessories second-pass audit: image clearly shows full-length cargo trousers; corrected to pants/trousers.'},
    '7578415118': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: image clearly shows white/black sneakers; corrected to shoes/sneakers.'},
    '7578513696': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies casual sneakers; corrected to shoes/sneakers.'},
    '7578503906': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies low-top skate sneakers; corrected to shoes/sneakers.'},
    '7578503936': {'category': 'shoe', 'subCategory': 'Sandals', 'review_note': 'Accessories second-pass audit: title clearly identifies casual slippers; corrected to shoes/sandals.'},
    '7576647651': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies casual sports shoes; corrected to shoes/sneakers.'},
    '7576554553': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies lightweight sports shoes; corrected to shoes/sneakers.'},
    '7576572323': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies arrow sports shoes; corrected to shoes/sneakers.'},
    '7576564481': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies lightweight casual sneakers; corrected to shoes/sneakers.'},
    '7762648301': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies casual sports shoes; corrected to shoes/sneakers.'},
    '7578420988': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies sports shoes; corrected to shoes/sneakers.'},
    '7576516991': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies breathable sports shoes; corrected to shoes/sneakers.'},
    '7578432628': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies breathable casual sneakers; corrected to shoes/sneakers.'},
    '7576597949': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies lightweight white sneakers; corrected to shoes/sneakers.'},
    '7578405434': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies running shoes; corrected to shoes/sneakers.'},
    '7778955647': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies platform sports shoes; corrected to shoes/sneakers.'},
    '7578444796': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies letter-print casual shoes; corrected to shoes/sneakers.'},
    '7576597951': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies fashion shoes; corrected to shoes/sneakers.'},
    '7576568443': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies board/sports shoes; corrected to shoes/sneakers.'},
    '7611861343': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies board/sports shoes; corrected to shoes/sneakers.'},
    '7778867065': {'category': 'shoe', 'subCategory': 'Sandals', 'review_note': 'Accessories second-pass audit: title clearly identifies casual slippers; corrected to shoes/sandals.'},
    '7578397536': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories second-pass audit: title clearly identifies men’s boxer briefs; corrected to clothing/underwear.'},
    '7578486682': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories second-pass audit: title clearly identifies men’s boxer briefs; corrected to clothing/underwear.'},
    '7576599907': {'category': 'clothing', 'subCategory': 'Jackets', 'review_note': 'Accessories second-pass audit: title clearly identifies a hooded down jacket; corrected to clothing/jackets.'},
    '7576550655': {'category': 'shoe', 'subCategory': 'Sandals', 'review_note': 'Accessories second-pass audit: title clearly identifies soft-bottom slippers; corrected to shoes/sandals.'},
    '7773159903': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: title clearly identifies lightweight sports shoes; corrected to shoes/sneakers.'},
    '7576550615': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: image clearly shows high-top sneakers; corrected to shoes/sneakers.'},
    '7611840843': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories second-pass audit: image clearly shows high-top sneakers; corrected to shoes/sneakers.'},
    '7543429795': {'category': 'clothing', 'subCategory': 'Sportswear', 'review_note': 'Accessories second-pass audit: image clearly shows coordinated zip jackets and trousers; corrected to clothing/sportswear.'},
    '7545244822': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories second-pass audit: image clearly shows men’s underwear; corrected to clothing/underwear.'},
    '7615163744': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories second-pass audit: image clearly shows men’s underwear; corrected to clothing/underwear.'},
    '7615126312': {'category': 'ACC', 'subCategory': 'Caps', 'review_note': 'Accessories second-pass audit: image clearly shows an embroidered knit hat; normalized to accessories/caps.'},
    '7762675943': {'category': 'clothing', 'subCategory': 'Sets', 'review_note': 'Accessories second-pass audit: image clearly shows a T-shirt and shorts assortment; corrected to clothing/sets.'},
    '7543416111': {'category': 'clothing', 'subCategory': 'Sportswear', 'review_note': 'Accessories second-pass audit: image clearly shows a coordinated track jacket, track pants, and cap; corrected to clothing/sportswear.'},
    '7543360659': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories second-pass audit: title clearly identifies underwear; corrected to clothing/underwear.'},
    '7783154995': {'category': 'clothing', 'subCategory': 'Sets', 'review_note': 'Accessories second-pass audit: image clearly shows a short-sleeve shirt and shorts set; corrected to clothing/sets.'},
    '7615095264': {'category': 'watches', 'subCategory': 'Watches', 'review_note': 'Accessories second-pass audit: title and image clearly show a luxury wristwatch; corrected to watches.'},
    '7614977634': {'category': 'shoe', 'subCategory': 'Sandals', 'review_note': 'Accessories second-pass audit: title and image clearly show casual clogs; corrected to shoes/sandals.'},
    '7554623854': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories second-pass audit: title and image clearly show underwear; corrected to clothing/underwear.'},
    '7612105467': {'category': 'clothing', 'subCategory': 'Underwear', 'review_note': 'Accessories second-pass audit: title clearly identifies men’s boxer briefs; corrected to clothing/underwear.'},
    '7612191743': {'category': 'watches', 'subCategory': 'Watches', 'review_note': 'Accessories second-pass audit: title clearly identifies a mechanical wristwatch; corrected to watches.'},
    '7794445183': {'category': 'ACC', 'subCategory': 'Caps', 'review_note': 'Accessories second-pass audit: title clearly identifies a fashion baseball cap; normalized to accessories/caps.'},
    '7612130579': {'category': 'ACC', 'subCategory': 'Caps', 'review_note': 'Accessories second-pass audit: title clearly identifies a breathable baseball cap; normalized to accessories/caps.'},
    '7554525338': {'category': 'ACC', 'subCategory': 'Caps', 'review_note': 'Accessories second-pass audit: title clearly identifies a hat; normalized to accessories/caps.'},
    '7612130605': {'category': 'watches', 'subCategory': 'Watches', 'review_note': 'Accessories second-pass audit: title clearly identifies a mechanical wristwatch; corrected to watches.'},
    '7578492088': {'category': 'shoe', 'subCategory': 'Sneakers', 'review_note': 'Accessories first-image audit: cover clearly shows high-top canvas shoes; corrected to shoes/sneakers.'},
    '7547272552': {'category': 'fragrance', 'subCategory': 'Perfume', 'review_note': 'Accessories first-image audit: cover clearly shows perfume packaging; corrected to fragrance/perfume.'},
    '7545264630': {'category': 'ACC', 'subCategory': 'Caps', 'review_note': 'Clothing first-image audit: cover clearly shows a cap assortment; corrected to accessories/caps.'},
    '7603647872': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'Clothing first-image audit: cover clearly shows full-length trousers; corrected to pants/trousers.'},
    '7601639507': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'Clothing first-image audit: cover clearly shows full-length trousers; corrected to pants/trousers.'},
    '7603561014': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'Clothing first-image audit: cover clearly shows full-length trousers; corrected to pants/trousers.'},
    '7601720125': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'Clothing first-image audit: cover clearly shows full-length trousers; corrected to pants/trousers.'},
    '7601649187': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'Clothing first-image audit: cover clearly shows denim jeans; corrected to pants/jeans.'},
    '7603578542': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'Clothing first-image audit: cover clearly shows denim jeans; corrected to pants/jeans.'},
    '7601661003': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'Clothing first-image audit: cover clearly shows full-length trousers; corrected to pants/trousers.'},
    '7601656693': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'Clothing first-image audit: cover clearly shows full-length trousers; corrected to pants/trousers.'},
    '7601717983': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'Clothing first-image audit: cover clearly shows a single jeans product; corrected to pants/jeans.'},
    '7603559462': {'category': 'pants', 'subCategory': 'Shorts', 'review_note': 'Clothing first-image audit: cover clearly shows shorts; corrected to pants/shorts.'},
    '7603632286': {'category': 'pants', 'subCategory': 'Trousers', 'review_note': 'Clothing first-image audit: cover clearly shows cargo trousers; corrected to pants/trousers.'},
    '7601661111': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'Clothing first-image audit: cover clearly shows denim jeans; corrected to pants/jeans.'},
    '7603600564': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'Clothing first-image audit: cover clearly shows denim jeans; corrected to pants/jeans.'},
    '7601680115': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'Clothing first-image audit: cover clearly shows denim jeans; corrected to pants/jeans.'},
    '7601680119': {'category': 'pants', 'subCategory': 'Jeans', 'review_note': 'Clothing first-image audit: cover clearly shows denim jeans; corrected to pants/jeans.'},
}


def classify_product(info: dict) -> str:
    title_text = f"{info.get('title', '')} {info.get('subcategory', '')}".lower()
    text = f"{title_text} {info.get('category', '')}".lower()
    if re.search(CATEGORY_PATTERNS['shoe'], title_text, re.I): return 'shoe'
    if re.search(CATEGORY_PATTERNS['watches'], title_text, re.I): return 'watches'
    if re.search(CATEGORY_PATTERNS['fragrance'], title_text, re.I): return 'fragrance'
    if re.search(CATEGORY_PATTERNS['bags'], title_text, re.I): return 'bags'
    if re.search(CATEGORY_PATTERNS['wallets'], title_text, re.I): return 'ACC'
    has_pants = bool(re.search(CATEGORY_PATTERNS['pants'], title_text, re.I))
    has_clothing = bool(re.search(CATEGORY_PATTERNS['clothing'], title_text, re.I))
    mixed_set = bool(re.search(r'\b(set|outfit|collection)\b', title_text, re.I))
    if has_pants and not (has_clothing and mixed_set): return 'pants'
    if has_clothing or info.get('category') == 'Clothing': return 'clothing'
    if re.search(CATEGORY_PATTERNS['accessories'], title_text, re.I) or info.get('category') == 'Accessories': return 'ACC'
    return 'ACC'

def subcategory_for(category: str, info: dict) -> str:
    text = f"{info.get('title', '')} {info.get('subcategory', '')}".lower()
    rules = {'pants': [('Jeans', r'jeans|denim'), ('Shorts', r'shorts'), ('Sweatpants', r'joggers|sweatpants|track pants'), ('Trousers', r'pants|trousers|cargo')], 'clothing': [('Hoodies', r'hoodies?|sweatshirts?'), ('Jackets', r'jackets?|coats?|outerwear'), ('Sweaters', r'sweaters?|knitwear'), ('Shirts', r'shirts?|tees?|t-?shirts?|jerseys?'), ('Dresses', r'dresses?|skirts?')], 'shoe': [('Sneakers', r'sneakers?|running shoes?'), ('Boots', r'boots?'), ('Sandals', r'sandals?|slides?')], 'bags': [('Backpacks', r'backpacks?'), ('Shoulder Bags', r'shoulder bags?|crossbody|handbags?')], 'fragrance': [('Perfume', r'perfumes?|fragrances?|colognes?|parfum')], 'watches': [('Watches', r'watch(?:es)?|腕表|手表')]}
    for label, pattern in rules.get(category, []):
        if re.search(pattern, text, re.I): return label
    return info.get('subcategory') if info.get('subcategory') and info.get('subcategory').lower() != 'unspecified' else 'Selection'

def brand_for(info: dict) -> str:
    text = f"{info.get('title', '')} {info.get('subcategory', '')}"
    for label, pattern in BRAND_PATTERNS:
        if re.search(pattern, text, re.I): return label
    return 'Unbranded'

def main() -> None:
    old_download = json.loads(DOWNLOAD_REPORT.read_text(encoding='utf-8'))
    image_by_url = {}
    for item in old_download.get('results', []):
        if item.get('ok') and item.get('url') and item.get('path'):
            stem = Path(item['path']).stem
            candidates = sorted((ROOT / 'client/public/product-images').glob(f'{stem}_*.webp'))
            if candidates:
                image_by_url[canon(item['url'])] = f'/product-images/{candidates[0].name}'
    cover_data = json.loads(COVER_LIST.read_text(encoding='utf-8'))
    cover_by_product = {str(item['product_id']): item.get('local_path') for item in cover_data['items'] if item.get('local_path')}

    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    products = {}
    sku_rows = defaultdict(list)
    image_rows = defaultdict(list)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration:
            continue
        headers = [str(v).strip() if v is not None else '' for v in header_values]
        index = {name: i for i, name in enumerate(headers)}
        for values in rows:
            def get(name: str):
                i = index.get(name)
                return values[i] if i is not None and i < len(values) else None
            pid = get('product_id')
            if pid is None:
                continue
            pid = str(pid).strip()
            if ws.title == 'products':
                raw_title = str(get('title_en_platform') or get('title_original') or '').strip()
                products[pid] = {
                    'title': clean_title(raw_title, f'Kakobuy Product {pid}'),
                    'title_original': str(get('title_original') or '').strip(),
                    'source_url': str(get('source_url') or '').strip(),
                    'platform_url': str(get('primary_platform_url') or '').strip(),
                    'category': str(get('category') or 'Unclassified').strip(),
                    'subcategory': str(get('subcategory') or '').strip(),
                    'seller': str(get('seller_name') or 'Kakobuy').strip(),
                    'collected_at': str(get('collected_at') or '').strip(),
                }
            elif ws.title == 'sku_records':
                sku_rows[pid].append({
                    'sku_id': str(get('sku_id') or '').strip(),
                    'variant': str(get('variant_label') or '').strip(),
                    'price_usd': money(get('price_usd')),
                    'price_rmb': money(get('price_rmb')),
                    'stock': str(get('stock_status') or get('status') or '').strip(),
                    'platform_url': str(get('platform_url') or '').strip(),
                    'checked_at': str(get('price_checked_at') or '').strip(),
                    'title': clean_title(str(get('title_en_platform') or get('title_original') or '').strip(), ''),
                    'title_original': str(get('title_original') or '').strip(),
                    'category': str(get('category') or '').strip(),
                    'subcategory': str(get('subcategory') or '').strip(),
                })
            elif ws.title == 'product_images':
                image_rows[pid].append({'url': str(get('image_url') or '').strip(), 'order': get('image_order') or 999})

    grouped = []
    for pid, rows in sku_rows.items():
        info = products.get(pid, {})
        fallback_row = next((row for row in rows if row.get('title')), {})
        if not info:
            info = {'title': fallback_row.get('title') or f'Kakobuy Product {pid}', 'category': fallback_row.get('category') or 'Unclassified', 'subcategory': fallback_row.get('subcategory') or '', 'seller': 'Kakobuy', 'source_url': '', 'platform_url': '', 'collected_at': ''}
        elif not info.get('title') or info.get('title') == f'Kakobuy Product {pid}':
            info['title'] = fallback_row.get('title') or info.get('title')
        prices = sorted({row['price_usd'] for row in rows if row['price_usd']})
        for price in prices:
            selected = [row for row in rows if row['price_usd'] == price]
            images = []
            if cover_by_product.get(pid):
                images.append(cover_by_product[pid])
            for image in sorted(image_rows.get(pid, []), key=lambda x: x['order']):
                local = image_by_url.get(canon(image['url']))
                if local and local not in images:
                    images.append(local)
            platform_url = next((row['platform_url'] for row in selected if row['platform_url']), info.get('platform_url', ''))
            collected_at = next((row['checked_at'] for row in selected if row['checked_at']), info.get('collected_at', ''))
            if not images:
                continue
            override = MANUAL_OVERRIDES.get(pid, {})
            suspected = SUSPECTED_REVIEW.get(pid, {})
            display_images = images[:16]
            remove_indices = {index for index in override.get('remove_image_indices', []) if isinstance(index, int)}
            display_images = [image for index, image in enumerate(display_images) if index not in remove_indices]
            custom_primary = override.get('primary_image_path')
            if isinstance(custom_primary, str) and custom_primary:
                display_images = [custom_primary] + [image for image in display_images if image != custom_primary]
            primary_index = override.get('primary_image_index')
            if isinstance(primary_index, int) and 0 <= primary_index < len(display_images):
                display_images = [display_images[primary_index]] + [image for index, image in enumerate(display_images) if index != primary_index]
            elif not remove_indices and not custom_primary:
                preferred_index = preferred_single_image(display_images)
                if preferred_index is not None:
                    display_images = [display_images[preferred_index]] + [image for index, image in enumerate(display_images) if index != preferred_index]
            inferred_category = classify_product(info)
            inferred_subcategory = subcategory_for(inferred_category, info)
            final_category = override.get('category') or inferred_category
            if str(final_category).strip().lower() == 'accessories':
                final_category = 'ACC'
            title_override = AI_TITLE_OVERRIDES.get(pid)
            display_title = finalize_display_title(title_override, final_category, override.get('subCategory') or inferred_subcategory) if title_override else optimize_display_title(info, final_category, override.get('subCategory') or inferred_subcategory, f'Kakobuy Product {pid}')
            grouped.append({
                'id': f'kb-{pid}-{price.replace(".", "-")}',
                'sourceProductId': pid,
                'name': display_title,
                'catalogName': display_title,
                'category': final_category,
                'subCategory': override.get('subCategory') or inferred_subcategory,
                'reviewStatus': 'suspected' if suspected else ('reviewed' if override else 'unreviewed'),
                'reviewNote': suspected.get('review_note') or override.get('review_note') or '',
                'brand': brand_for(info),
                'price': float(price),
                'referencePrice': float(price),
                'currency': 'USD',
                'description': 'Product details are based on the latest Kakobuy catalog capture.',
                'sizes': sorted({row['variant'] for row in selected if row['variant']}),
                'colors': [],
                'stock': 'In stock' if any('available' in row['stock'].lower() for row in selected) else 'Check availability',
                'shop': info.get('seller') or 'Kakobuy',
                'shopUrl': info.get('source_url', ''),
                'url': platform_url,
                'platformLinks': platform_links_for(pid),
                'images': display_images,
                'tags': [value for value in ['kakobuy', subcategory_for(classify_product(info), info), classify_product(info)] if value],
                'collectedAt': collected_at,
                'sourceSkuIds': [row['sku_id'] for row in selected if row['sku_id']],
                'priceRmb': next((float(row['price_rmb']) for row in selected if row['price_rmb']), None),
                'priceCheckedAt': collected_at,
            })

    grouped.sort(key=lambda item: (item['category'], item['catalogName'].lower(), item['price'], item['id']))
    OUTPUT_JSON.write_text(json.dumps(grouped, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    header = 'export type Product = { id: string; name: string; catalogName: string; category: string; subCategory: string; reviewStatus?: "reviewed" | "suspected" | "unreviewed"; reviewNote?: string; brand: string; price: number; referencePrice: number | null; currency: string; description: string; sizes: string[]; colors: string[]; stock: string; shop: string; shopUrl: string; url: string; platformLinks?: Record<string, string>; images: string[]; tags: string[]; collectedAt: string; sourceProductId?: string; sourceSkuIds?: string[];\n priceRmb?: number | null; priceCheckedAt?: string }\n'
    body = 'export const products: Product[] = ' + json.dumps(grouped, ensure_ascii=False, indent=2) + ' as Product[];\n'
    footer = 'export const categoryLabels: Record<string, string> = { clothing: "Clothing", shoe: "Shoes", pants: "Pants", bags: "Bags", fragrance: "Fragrance", ACC: "Accessories", watches: "Watches" };\nexport const categoryOrder = ["all", "clothing", "pants", "shoe", "bags", "fragrance", "watches", "ACC"];\n'
    OUTPUT_TS.write_text(header + body + footer, encoding='utf-8')
    print(f'Generated {len(grouped)} grouped Kakobuy products with local images; unique source products: {len(sku_rows)}')

if __name__ == '__main__':
    main()
