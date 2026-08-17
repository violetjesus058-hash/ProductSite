from openpyxl import load_workbook
from pathlib import Path
import json

workbook = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
wb = load_workbook(workbook, read_only=True, data_only=True)
result = {}
for ws in wb.worksheets:
    rows = ws.iter_rows(values_only=True)
    headers = [str(x or '').strip() for x in next(rows)]
    result[ws.title] = {'headers': headers, 'samples': []}
    wanted = {'7543313315','7543429795','7601654481','7576550615','7612130605'}
    index = {h:i for i,h in enumerate(headers)}
    for row in rows:
        row_dict = {h: row[i] if i < len(row) else None for h,i in index.items()}
        pid = str(row_dict.get('source_product_id') or row_dict.get('product_id') or row_dict.get('item_id') or '')
        if pid in wanted and len(result[ws.title]['samples']) < 20:
            result[ws.title]['samples'].append(row_dict)
print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
