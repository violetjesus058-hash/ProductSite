import json
from collections import Counter
from pathlib import Path
from openpyxl import load_workbook
root=Path('/home/ubuntu/product-catalog-site')
text=(root/'client/src/data/products.ts').read_text()
start=text.index('export const products: Product[] = ')+len('export const products: Product[] = ')
end=text.index('] as Product[];',start)+1
products=json.loads(text[start:end])
target={str(p['sourceProductId']) for p in products if p.get('catalogName')=='Essential Apparel'}
wb=load_workbook('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx',read_only=True,data_only=True)
ws=wb.active
it=ws.iter_rows(values_only=True); header=list(next(it)); idx={v:i for i,v in enumerate(header)}
seen={}
for row in it:
 pid=str(row[idx['product_id']]) if row[idx['product_id']] is not None else ''
 if pid in target and pid not in seen:
  seen[pid]=(row[idx['title_en_platform']],row[idx['title_original']],row[idx['subcategory']])
print('target products',len(target),'source matched',len(seen))
for key,count in Counter(v[0] for v in seen.values()).most_common(80): print(count, key)
