from openpyxl import load_workbook
from pathlib import Path
import json, re

workbook = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
wb = load_workbook(workbook, read_only=True, data_only=True)
pattern = re.compile(r'(ALL[-_ ]?[A-Z]+[-_ ]?\d+|High[- ]?Quality(?: [A-Z]+)? \d+[-_A-Z]*|High-quality \d+[-_A-Z]*|REP High Quality \d+[-_A-Z]*|\d+[-_][A-Z]{1,5}[-_]\d+)', re.I)
rows_out = []
for ws in wb.worksheets:
    if ws.title != 'products':
        continue
    rows = ws.iter_rows(values_only=True)
    headers = [str(x or '').strip() for x in next(rows)]
    index = {h:i for i,h in enumerate(headers)}
    for row in rows:
        d = {h: row[i] if i < len(row) else None for h,i in index.items()}
        title = str(d.get('title_en_platform') or d.get('title_original') or '').strip()
        if pattern.search(title):
            rows_out.append({k:d.get(k) for k in ['product_id','title_original','title_en_platform','category','subcategory','seller_name','source_sheets','default_price_usd']})
print(json.dumps(rows_out, ensure_ascii=False, indent=2, default=str))
