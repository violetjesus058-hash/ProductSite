from __future__ import annotations

from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
REPORT = Path('/home/ubuntu/product-catalog-site/scripts/kakobuy-price-grouping-analysis.md')


def money(value: object) -> str:
    try:
        return f'{Decimal(str(value)).quantize(Decimal("0.01"))}'
    except (InvalidOperation, TypeError, ValueError):
        return ''


def main() -> None:
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    titles = {}
    categories = {}
    product_rows = defaultdict(list)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration:
            continue
        headers = [str(v).strip() if v is not None else '' for v in header_values]
        index = {header: i for i, header in enumerate(headers)}
        if ws.title not in {'products', 'sku_records'}:
            continue
        for values in rows:
            def get(name: str):
                i = index.get(name)
                return values[i] if i is not None and i < len(values) else None
            pid = get('product_id')
            if pid is None:
                continue
            pid = str(pid).strip()
            if ws.title == 'products':
                titles[pid] = str(get('title_en_platform') or get('title_original') or '').strip()
                categories[pid] = str(get('category') or 'Unclassified').strip()
            else:
                product_rows[pid].append({
                    'sku_id': str(get('sku_id') or '').strip(),
                    'variant': str(get('variant_label') or '').strip(),
                    'price_usd': money(get('price_usd')),
                    'price_rmb': money(get('price_rmb')),
                    'status': str(get('status') or '').strip(),
                    'stock_status': str(get('stock_status') or '').strip(),
                    'stock_quantity': get('stock_quantity'),
                })

    group_counts = {}
    all_price_groups = 0
    one_price_products = []
    multi_price_products = []
    no_price_products = []
    price_distribution = Counter()
    category_distribution = Counter()
    samples = []
    for pid, rows in product_rows.items():
        prices = sorted({row['price_usd'] for row in rows if row['price_usd']})
        group_counts[pid] = len(prices)
        all_price_groups += len(prices)
        price_distribution[len(prices)] += 1
        category_distribution[categories.get(pid, 'Unclassified')] += 1
        if len(prices) == 0:
            no_price_products.append(pid)
        elif len(prices) == 1:
            one_price_products.append(pid)
        else:
            multi_price_products.append(pid)
            if len(samples) < 30:
                by_price = {price: [r['variant'] for r in rows if r['price_usd'] == price][:6] for price in prices}
                samples.append((pid, titles.get(pid, ''), categories.get(pid, ''), prices, by_price))

    report = [
        '# Kakobuy 按价格聚合 SKU 分析', '',
        '> 聚合规则：以 `product_id + price_usd` 作为一个展示商品组。同一 `product_id` 下所有可售型号的美元价格只有一个时，合并为一个商品；存在多个美元价格时，拆分为多个价格商品。SKU 的款式、尺码和库存保留在该价格组内部。', '',
        '| 指标 | 数量 |', '|---|---:|',
        f'| Kakobuy 商品数 | {len(product_rows)} |',
        f'| SKU 明细数 | {sum(len(rows) for rows in product_rows.values())} |',
        f'| 无美元价格商品 | {len(no_price_products)} |',
        f'| 全部型号同一美元价格，可合并 | {len(one_price_products)} |',
        f'| 存在多个美元价格，需要拆分 | {len(multi_price_products)} |',
        f'| 按商品+美元价格聚合后的展示商品数 | {all_price_groups} |',
        f'| 理论上比按 SKU 行展示减少 | {sum(len(rows) for rows in product_rows.values()) - all_price_groups} |',
        '', '## 每个商品价格组数量分布', '', '| 价格组数量 | 商品数 |', '|---:|---:|',
    ]
    for group_count, product_count in sorted(price_distribution.items()):
        report.append(f'| {group_count} | {product_count} |')

    report.extend(['', '## 按类别统计', '', '| Category | 商品数 |', '|---|---:|'])
    for category, count in category_distribution.most_common():
        report.append(f'| {category} | {count} |')

    report.extend(['', '## 多价格商品示例', '', '| Product ID | Title | Category | USD prices | Example variants |', '|---|---|---|---|---|'])
    for pid, title, category, prices, by_price in samples:
        examples = '; '.join(f'{price}: {", ".join(labels) or "(variant labels unavailable)"}' for price, labels in by_price.items())
        report.append(f'| {pid} | {title[:80].replace("|", "/")} | {category} | {", ".join(prices)} | {examples[:180].replace("|", "/")} |')

    report.extend(['', '## 实施建议', '', '价格组应生成独立的展示记录，但不需要复制所有图片。每个 `product_id + price_usd` 组共享同一商品的图片池；只有当不同价格组的图片明确不同，才按 SKU 选项图片做分组。商品详情页显示该价格组的 USD 价格，并在规格选择中只列出属于该价格组的型号。', '', '对于同一商品所有型号价格相同的记录，保留一个商品卡片和一个主价格，型号、尺码、颜色作为选项，不再生成重复卡片。对于价格不同的记录，使用稳定的派生 ID，例如 `kakobuy-{product_id}-{price_usd}` 的哈希形式，避免因排序变化导致详情页链接漂移。', '', '图片下载应从 `product_images` 按 `product_id` 聚合，并先复用现有托管图片；不要从 77,563 条 `sku_records` 重复下载。', ''])
    REPORT.write_text('\n'.join(report) + '\n', encoding='utf-8')
    print(f'Wrote {REPORT}')
    print(f'Products: {len(product_rows)}; SKU rows: {sum(len(rows) for rows in product_rows.values())}; one-price: {len(one_price_products)}; multi-price: {len(multi_price_products)}; grouped displays: {all_price_groups}; no-price: {len(no_price_products)}')


if __name__ == '__main__':
    main()
