from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook

WORKBOOK = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
PRODUCTS = Path('/home/ubuntu/product-catalog-site/client/src/data/products.ts')
REPORT = Path('/home/ubuntu/product-catalog-site/scripts/sku-workbook-analysis.md')


def normalize(value: object) -> str:
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip().lower()


def extract_products() -> tuple[list[dict], str | None]:
    text = PRODUCTS.read_text(encoding='utf-8')
    match = re.search(r'export const products: Product\[\] = (\[.*?\]) as Product\[\];', text, re.S)
    if not match:
        return [], 'products array marker was not found'
    try:
        return json.loads(match.group(1)), None
    except json.JSONDecodeError as exc:
        return [], f'JSON parse failed at line {text[:exc.pos].count(chr(10)) + 1}, column {exc.colno}: {exc.msg}'


def pick(record: dict, candidates: list[str]) -> object:
    normalized = {normalize(k): v for k, v in record.items()}
    for candidate in candidates:
        key = normalize(candidate)
        if key in normalized:
            return normalized[key]
    for key, value in normalized.items():
        if any(normalize(candidate) in key for candidate in candidates):
            return value
    return None


def main() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    products, parse_error = extract_products()
    product_by_id = {str(item.get('id')): item for item in products if item.get('id') is not None}
    existing_ids = set(product_by_id)

    id_fields = ['商品ID', '商品 id', 'itemid', 'item id', 'id', '商品编号', '货号']
    title_fields = ['商品名称', '产品名称', 'title', 'name', '商品标题', '产品标题']
    price_fields = ['价格', 'price', '售价', '商品价格', 'sku价格', 'sku price']
    url_fields = ['商品链接', '商品网址', 'url', 'link', 'fansbuy链接', 'kakobuy链接']

    sheet_stats = []
    field_counts = Counter()
    field_samples: defaultdict[str, list[str]] = defaultdict(list)
    rows_with_ids: list[tuple[str, dict]] = []
    total_rows = 0

    for ws in workbook.worksheets:
        iterator = ws.iter_rows(values_only=True)
        try:
            header_values = next(iterator)
        except StopIteration:
            sheet_stats.append((ws.title, 0, 0, []))
            continue
        headers = [str(value).strip() if value is not None else '' for value in header_values]
        sheet_rows = 0
        for row_number, values in enumerate(iterator, start=2):
            sheet_rows += 1
            total_rows += 1
            record = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}
            record['_sheet'] = ws.title
            record['_row'] = row_number
            for key, value in record.items():
                if key.startswith('_') or value in (None, ''):
                    continue
                field_counts[key] += 1
                if len(field_samples[key]) < 3:
                    sample = str(value).replace('\n', ' ')[:120]
                    if sample not in field_samples[key]:
                        field_samples[key].append(sample)
            item_id = pick(record, id_fields)
            if item_id is not None and str(item_id).strip():
                rows_with_ids.append((str(item_id).strip(), record))
        sheet_stats.append((ws.title, sheet_rows, len(headers), headers))

    workbook_ids = [item_id for item_id, _ in rows_with_ids]
    duplicate_ids = [item_id for item_id, count in Counter(workbook_ids).items() if count > 1]
    overlap_ids = sorted(set(workbook_ids) & existing_ids)
    new_ids = sorted(set(workbook_ids) - existing_ids)
    missing_ids = sorted(existing_ids - set(workbook_ids))

    sections = [
        '# Latest Fansbuy SKU Workbook Analysis', '',
        f'- Workbook: `{WORKBOOK.name}`',
        f'- Existing catalog source: `{PRODUCTS.name}`', '',
        '## Workbook overview', '',
        '| Sheet | Data rows | Columns | Headers |', '|---|---:|---:|---|',
    ]
    for title, row_count, column_count, headers in sheet_stats:
        sections.append(f"| {title} | {row_count} | {column_count} | {'; '.join(h or '[blank]' for h in headers).replace('|', '/')} |")

    sections.extend(['', '## Existing catalog parse status', '', f'- Existing product objects parsed: **{len(products)}**'])
    sections.append(f'- Parse issue: **{parse_error}**' if parse_error else '- Parse issue: none')

    sections.extend(['', '## Workbook fields', '', '| Field | Non-empty rows | Example values |', '|---|---:|---|'])
    for key, count in field_counts.most_common():
        sections.append(f"| {key} | {count} | {'; '.join(field_samples[key]).replace('|', '/')} |")

    sections.extend(['', '## ID and overlap analysis', '', '| Metric | Value |', '|---|---:|'])
    metrics = [
        ('Workbook data rows', total_rows),
        ('Rows with a detectable product/SKU ID', len(rows_with_ids)),
        ('Unique workbook IDs', len(set(workbook_ids))),
        ('Duplicate workbook IDs', len(duplicate_ids)),
        ('IDs overlapping current catalog', len(overlap_ids)),
        ('IDs new to current catalog', len(new_ids)),
        ('Current catalog IDs absent from workbook', len(missing_ids)),
    ]
    for label, value in metrics:
        sections.append(f'| {label} | {value} |')
    if duplicate_ids:
        sections.extend(['', f"Duplicate IDs: `{', '.join(duplicate_ids[:50])}`"])

    sections.extend(['', '## Sample normalized records', '', '| Sheet row | ID | Title | Price | URL |', '|---:|---|---|---|---|'])
    for item_id, record in rows_with_ids[:30]:
        title = pick(record, title_fields)
        price = pick(record, price_fields)
        url = pick(record, url_fields)
        sections.append(f"| {record['_sheet']}:{record['_row']} | {item_id} | {str(title or '')[:90].replace('|', '/')} | {str(price or '')[:40].replace('|', '/')} | {str(url or '')[:100].replace('|', '/')} |")

    # Fuzzy matching is intentionally capped; ID matching remains authoritative.
    fuzzy_matches = []
    catalog_titles = [(existing_id, normalize(product.get('name') or product.get('catalogName'))) for existing_id, product in product_by_id.items()]
    seen_titles = set()
    for item_id, record in rows_with_ids:
        if item_id in existing_ids:
            continue
        title = normalize(pick(record, title_fields))
        if not title or title in seen_titles or len(fuzzy_matches) >= 100:
            continue
        seen_titles.add(title)
        best_score, best_id, best_title = 0.0, None, ''
        for existing_id, existing_title in catalog_titles:
            if not existing_title:
                continue
            score = SequenceMatcher(None, title, existing_title).ratio()
            if score > best_score:
                best_score, best_id, best_title = score, existing_id, existing_title
        if best_score >= 0.72:
            fuzzy_matches.append((best_score, item_id, title, best_id, best_title))
    fuzzy_matches.sort(reverse=True)

    sections.extend(['', '## Fuzzy title matches for non-overlapping IDs', '', '| Similarity | Workbook ID | Workbook title | Existing ID | Existing title |', '|---:|---|---|---|---|'])
    if fuzzy_matches:
        for score, workbook_id, workbook_title, existing_id, existing_title in fuzzy_matches:
            sections.append(f'| {score:.3f} | {workbook_id} | {workbook_title[:80].replace("|", "/")} | {existing_id} | {existing_title[:80].replace("|", "/")} |')
    else:
        sections.append('| — | No high-confidence fuzzy title matches found | | | |')

    sections.extend([
        '', '## Recommended update policy', '',
        '> Do not replace the current catalog in place until the workbook fields are mapped and every row has a stable identity.', '',
        '1. Treat a stable product/item ID as the primary key. Treat a SKU option ID as a child record, not as a separate product card, unless the product page itself has no stable item ID.',
        '2. For overlapping IDs, update price and SKU-level option data only after validating the workbook currency and price semantics. Preserve the current image URLs, editorial catalog name, category, and platform links unless the workbook explicitly provides a newer verified value.',
        '3. For new IDs, stage them in an import report first. Do not automatically publish them until image availability, title, category, and Fansbuy URL are complete.',
        '4. For duplicate IDs, group all rows under one product and retain one price per SKU option. Never silently choose the first duplicate row.',
        '5. Write an import snapshot and a machine-readable mapping report before changing `products.ts`, so the operation can be rolled back without reconstructing data from the workbook.',
        '',
    ])

    REPORT.write_text('\n'.join(sections) + '\n', encoding='utf-8')
    print(f'Wrote {REPORT}')
    print(f'Workbook sheets: {len(workbook.worksheets)}; rows: {total_rows}; IDs: {len(set(workbook_ids))}; overlap: {len(overlap_ids)}; new: {len(new_ids)}; missing: {len(missing_ids)}; duplicate IDs: {len(duplicate_ids)}')


if __name__ == '__main__':
    main()
