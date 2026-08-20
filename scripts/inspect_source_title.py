from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
wb = load_workbook(path, read_only=True, data_only=True)
for ws in wb.worksheets:
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(v).strip() if v is not None else '' for v in next(rows)]
    except StopIteration:
        continue
    if 'product_id' not in headers:
        continue
    indexes = {name: index for index, name in enumerate(headers)}
    for values in rows:
        pid = str(values[indexes['product_id']]) if indexes['product_id'] < len(values) and values[indexes['product_id']] is not None else ''
        if pid in {'7612152297', '7554543542'}:
            result = {key: values[index] if index < len(values) else None for key, index in indexes.items() if key in {'product_id', 'title_en_platform', 'title_original'}}
            print(repr(result))
