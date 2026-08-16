from __future__ import annotations

import json
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook

WORKBOOK = Path('/home/ubuntu/upload/pasted_file_drov2V_Kakobuy_SKU_数据表.xlsx')
DOWNLOAD_REPORT = Path('/home/ubuntu/product-catalog-site/scripts/product-image-download-report.json')
LOCAL_IMAGE_DIR = Path('/home/ubuntu/product-catalog-site/client/public/product-images')
OUTPUT_JSON = Path('/home/ubuntu/product-catalog-site/scripts/kakobuy-sample-products.json')
OUTPUT_MD = Path('/home/ubuntu/product-catalog-site/scripts/kakobuy-sample-products.md')


def money(value: object) -> str:
    try:
        return f'{Decimal(str(value)).quantize(Decimal("0.01"))}'
    except (InvalidOperation, TypeError, ValueError):
        return ''


def canon(url: object) -> str:
    if not url:
        return ''
    return re.sub(r'\?.*$', '', str(url).strip()).lower()


def main() -> None:
    download = json.loads(DOWNLOAD_REPORT.read_text(encoding='utf-8'))
    existing_map = {}
    for item in download.get('results', []):
        if not item.get('ok') or not item.get('url') or not item.get('path'):
            continue
        source_stem = Path(item['path']).stem
        candidates = sorted(LOCAL_IMAGE_DIR.glob(f'{source_stem}_*.webp'))
        if candidates:
            existing_map[canon(item['url'])] = f'/product-images/{candidates[0].name}'
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    products = {}
    sku_rows = defaultdict(list)
    image_rows = defaultdict(list)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            header_values = next(rows)
        except StopIteration:
            continue
        headers = [str(v).strip() if v is not None else '' for v in header_values]
        index = {name: i for i, name in enumerate(headers)}
        for values in rows:
            def get(name: str):
                i = index.get(name)
                return values[i] if i is not None and i < len(values) else None
            pid = get('product_id')
            if pid is None:
                continue
            pid = str(pid).strip()
            if ws.title == 'products':
                products[pid] = {
                    'title': str(get('title_en_platform') or get('title_original') or f'Kakobuy Product {pid}').strip(),
                    'original_title': str(get('title_original') or '').strip(),
                    'source_url': str(get('source_url') or '').strip(),
                    'platform_url': str(get('primary_platform_url') or '').strip(),
                    'category': str(get('category') or 'Unclassified').strip(),
                    'subcategory': str(get('subcategory') or '').strip(),
                    'seller': str(get('seller_name') or 'Kakobuy').strip(),
                    'collected_at': str(get('collected_at') or '').strip(),
                }
            elif ws.title == 'sku_records':
                sku_rows[pid].append({
                    'sku_id': str(get('sku_id') or '').strip(),
                    'variant': str(get('variant_label') or '').strip(),
                    'price_usd': money(get('price_usd')),
                    'price_rmb': money(get('price_rmb')),
                    'stock': str(get('stock_status') or get('status') or '').strip(),
                    'quantity': get('stock_quantity'),
                    'platform_url': str(get('platform_url') or '').strip(),
                })
            elif ws.title == 'product_images':
                image_rows[pid].append({
                    'url': str(get('image_url') or '').strip(),
                    'type': str(get('image_type') or '').strip(),
                    'order': get('image_order') or 999,
                })

    grouped = []
    for pid, rows in sku_rows.items():
        info = products.get(pid, {})
        prices = sorted({row['price_usd'] for row in rows if row['price_usd']})
        for price in prices:
            selected = [row for row in rows if row['price_usd'] == price]
            image_paths = []
            for image in sorted(image_rows.get(pid, []), key=lambda x: x['order']):
                path = existing_map.get(canon(image['url']))
                if path and path not in image_paths:
                    image_paths.append(path)
            platform_url = next((row['platform_url'] for row in selected if row['platform_url']), info.get('platform_url', ''))
            product = {
                'id': f'kb-{pid}-{price.replace(".", "-")}',
                'sourceProductId': pid,
                'name': info.get('title', f'Kakobuy Product {pid}'),
                'catalogName': info.get('title', f'Kakobuy Product {pid}'),
                'category': {'Clothing': 'clothing', 'Shoes': 'shoe', 'Accessories': 'ACC'}.get(info.get('category'), 'ACC'),
                'subCategory': info.get('subcategory') or 'Selection',
                'brand': 'Unbranded',
                'price': float(price),
                'referencePrice': float(price),
                'currency': 'USD',
                'description': 'Product details are based on the latest Kakobuy catalog capture.',
                'sizes': sorted({row['variant'] for row in selected if row['variant']}),
                'colors': [],
                'stock': 'In stock' if any('available' in row['stock'].lower() for row in selected) else 'Check availability',
                'shop': info.get('seller') or 'Kakobuy',
                'shopUrl': info.get('source_url', ''),
                'url': platform_url,
                'images': image_paths,
                'tags': [value for value in ['kakobuy', info.get('subcategory', ''), info.get('category', '')] if value],
                'collectedAt': info.get('collected_at', ''),
                'sourceSkuIds': [row['sku_id'] for row in selected if row['sku_id']],
                'priceRmb': next((float(row['price_rmb']) for row in selected if row['price_rmb']), None),
                'priceCheckedAt': info.get('collected_at', ''),
            }
            grouped.append(product)

    grouped.sort(key=lambda item: item['id'])
    sample = [item for item in grouped if item['images'] and item['url']][:10]
    OUTPUT_JSON.write_text(json.dumps(sample, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    lines = ['# Kakobuy 样本商品', '', f'- 聚合总商品组：{len(grouped)}', f'- 样本商品数：{len(sample)}', f'- 样本中复用现有图片的商品数：{sum(bool(item["images"]) for item in sample)}', '', '| ID | Title | USD | Variants | Images | Kakobuy URL |', '|---|---|---:|---:|---:|---|']
    for item in sample:
        lines.append(f"| {item['id']} | {item['name'][:70].replace('|', '/')} | ${item['price']:.2f} | {len(item['sizes'])} | {len(item['images'])} | {item['url'][:110].replace('|', '/')} |")
    OUTPUT_MD.write_text('\n'.join(lines) + '\n', encoding='utf-8')
    print(f'Grouped products: {len(grouped)}; sample products: {len(sample)}; sample output: {OUTPUT_JSON}')


if __name__ == '__main__':
    main()
